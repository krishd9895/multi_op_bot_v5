import os
import logging
import random
import calendar
from datetime import datetime, timedelta
from typing import List, Dict
import pandas as pd
import pymongo
from pymongo import MongoClient
import pytz
import telebot
from telebot import types
import schedule
import time
import threading
from io import BytesIO
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from calendar import monthrange
from datetime import date as dt_date
from webserver import keep_alive

# Custom file handler that limits lines to 6000
class LimitedLinesFileHandler(logging.FileHandler):
    def __init__(self, filename, max_lines=6000, mode='a', encoding=None, delay=False):
        super().__init__(filename, mode, encoding, delay)
        self.max_lines = max_lines
        self.filename = filename
        self.last_schedule_check = None
        self._check_and_rotate()

    def _check_and_rotate(self):
        """Check if file exceeds max_lines and rotate if needed"""
        try:
            with open(self.filename, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            if len(lines) > self.max_lines:
                # Keep only the last max_lines lines
                lines_to_keep = lines[-self.max_lines:]
                with open(self.filename, 'w', encoding='utf-8') as f:
                    f.writelines(lines_to_keep)
                print(f"Log file rotated: kept last {self.max_lines} lines")
                # Add a rotation marker to the log file
                with open(self.filename, 'a', encoding='utf-8') as f:
                    f.write(f"{datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')} - LOG_ROTATION - INFO - Log file rotated, kept last {self.max_lines} lines\n")
        except FileNotFoundError:
            # File doesn't exist yet, that's fine
            pass
        except Exception as e:
            print(f"Error rotating log file: {e}")

    def emit(self, record):
        # Special handling for schedule check logs
        if record.msg.startswith('🕰️ Schedule check'):
            try:
                with open(self.filename, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                # Filter out old schedule check logs while keeping other logs
                filtered_lines = []
                for line in lines:
                    # Keep the line if it's not a schedule check log
                    if ' - DEBUG - 🕰️ Schedule check at ' not in line:
                        filtered_lines.append(line)
                # Write filtered lines back to file
                with open(self.filename, 'w', encoding='utf-8') as f:
                    f.writelines(filtered_lines)
            except Exception as e:
                print(f"Error managing schedule check logs: {e}")

        # Check rotation before each emit
        self._check_and_rotate()
        super().emit(record)

# Configure logging
logger = logging.getLogger(__name__)

# Configuration
BOT_TOKEN = os.getenv('BOT_TOKEN')
MONGODB_URI = os.getenv('MONGODB_URI')
DB_NAME = os.getenv('DB_NAME', 'TD')
OWNER_ID = os.getenv('OWNER_ID') # Replace with the actual owner user_id
USER_TIMEOUT = 60

# Daily schedule times (IST)
DAILY_PROMPT_TIME = "19:00"  # 7:00 PM IST - Daily activity prompt
DEFAULT_ACTIVITY_TIME = "20:00"  # 8:00 PM IST - Default activity fallback

# Global month-wise main activities
MAIN_ACTIVITIES_BY_MONTH = {
    1: ["Attended this village to observe sowing operations", "Attended this village to observe seasonal and crop conditions"],
    2: ["Attended this village to observe sowing operations", "Attended this village to observe seasonal and crop conditions"],
    3: ["Attended this village to observe seasonal and crop conditions"],
    4: ["Attended this village to observe harvesting operations", "Attended this village to observe seasonal and crop conditions"],
    5: ["Attended this village to observe harvesting operations", "Attended this village to observe seasonal and crop conditions"],
    6: ["Attended this village to observe seasonal and crop conditions"],
    7: ["Attended this village to observe sowing operations", "Attended this village to observe seasonal and crop conditions"],
    8: ["Attended this village to observe sowing operations", "Attended this village to observe seasonal and crop conditions"],
    9: ["Attended this village to observe seasonal and crop conditions"],
    10: ["Attended this village to observe sowing operations", "Attended this village to observe harvesting operations", "Attended this village to observe seasonal and crop conditions"],
    11: ["Attended this village to observe sowing operations", "Attended this village to observe seasonal and crop conditions"],
    12: ["Attended this village to observe seasonal and crop conditions"]
}


# MongoDB setup
client = MongoClient(MONGODB_URI)
db = client[DB_NAME]
users_collection = db.users
main_activities_collection = db.main_activities
config_collection = db.config

# Timezone
IST = pytz.timezone('Asia/Kolkata')

def load_schedule_times():
    """Load schedule times from MongoDB, or use defaults."""
    global DAILY_PROMPT_TIME, DEFAULT_ACTIVITY_TIME
    try:
        config = config_collection.find_one({'_id': 'schedule_times'})
        if config:
            DAILY_PROMPT_TIME = config.get('daily_prompt_time', "19:00")
            DEFAULT_ACTIVITY_TIME = config.get('default_activity_time', "20:00")
            logger.info(f"Loaded schedule times from DB: Prompt at {DAILY_PROMPT_TIME}, Fallback at {DEFAULT_ACTIVITY_TIME}")
        else:
            logger.info(f"No schedule times in DB, using default values: Prompt at {DAILY_PROMPT_TIME}, Fallback at {DEFAULT_ACTIVITY_TIME}")
    except Exception as e:
        logger.error(f"Error loading schedule times from DB: {e}. Using default values.")


# --- Logging setup ---
LOG_FILENAME = 'logs.txt'

# Delete previous log file on bot restart
try:
    if os.path.exists(LOG_FILENAME):
        os.remove(LOG_FILENAME)
        print(f"Deleted previous log file: {LOG_FILENAME}")
except Exception as e:
    print(f"Error deleting previous log file: {e}")

logger.setLevel(logging.DEBUG)

# Remove any existing handlers to avoid duplicates
for handler in logger.handlers[:]:
    logger.removeHandler(handler)

# Use custom handler that limits to 6000 lines
file_handler = LimitedLinesFileHandler(LOG_FILENAME, max_lines=6000, encoding='utf-8')
file_handler.setLevel(logging.DEBUG)

# Create formatter with IST timezone
class ISTFormatter(logging.Formatter):
    def formatTime(self, record, datefmt=None):
        dt = datetime.fromtimestamp(record.created, tz=IST)
        if datefmt:
            return dt.strftime(datefmt)
        else:
            return dt.strftime('%Y-%m-%d %H:%M:%S IST')

formatter = ISTFormatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# Console handler removed to prevent duplicate logs

class TourDiaryBot:
    def __init__(self):
        self.bot = telebot.TeleBot(BOT_TOKEN)
        self.callback_data = {}
        self.cancelled_users = set()  # Track users who cancelled input
        self.input_prompt_message = {}  # Track prompt message_id per user
        self.pending_prompts = {}  # Track pending prompts for timeout
        self.daily_prompt_message_ids = {}  # Store daily prompt message IDs for deletion
        load_schedule_times()
        self.setup_handlers()
        self.schedule_daily_tasks()
        self.start_prompt_timeout_checker()

        # Log bot startup with scheduled times
        current_time = datetime.now(IST)
        logger.info(f"🚀 TD Bot started at {current_time.strftime('%Y-%m-%d %H:%M:%S IST')}")
        logger.info(f"📅 Scheduled times: Daily prompt at {DAILY_PROMPT_TIME} IST, Default activity at {DEFAULT_ACTIVITY_TIME} IST")
        logger.info(f"🔍 Variable values - DAILY_PROMPT_TIME: '{DAILY_PROMPT_TIME}' (type: {type(DAILY_PROMPT_TIME)})")
        logger.info(f"🔍 Variable values - DEFAULT_ACTIVITY_TIME: '{DEFAULT_ACTIVITY_TIME}' (type: {type(DEFAULT_ACTIVITY_TIME)})")

    def start_prompt_timeout_checker(self):
        def check_timeouts():
            while True:
                now = time.time()
                to_remove = []
                for user_id, prompt in list(self.pending_prompts.items()):
                    if now >= prompt['timeout_time']:
                        try:
                            self.bot.delete_message(prompt['chat_id'], prompt['message_id'])
                        except Exception as e:
                            logger.error(f"Error deleting timed out prompt for user {user_id}: {e}")
                        # Clean up any pending state
                        if user_id in self.callback_data:
                            del self.callback_data[user_id]
                        if user_id in self.input_prompt_message:
                            del self.input_prompt_message[user_id]
                        self.cancelled_users.add(user_id)
                        try:
                            self.bot.send_message(prompt['chat_id'], "⏰ Timed out. Operation cancelled.")
                        except Exception as e:
                            logger.error(f"Error sending timeout message to user {user_id}: {e}")
                        to_remove.append(user_id)
                for user_id in to_remove:
                    del self.pending_prompts[user_id]
                time.sleep(2)
        thread = threading.Thread(target=check_timeouts, daemon=True)
        thread.start()

    def start_command(self, message):
        """Handle /start command"""
        user_id = message.from_user.id
        user = users_collection.find_one({'user_id': user_id})

        if not user:
            users_collection.insert_one({
                'user_id': user_id,
                'headquarters': None,
                'villages': [],
                'activities': [],
                'custom_activities': [],
                'role': None
            })
        message_text = "🎉 Welcome to TD Bot!"
        self.bot.reply_to(message, message_text)

    def handle_file_upload(self, message):
        user_id = message.from_user.id
        file = message.document
        if not file:
            return
        file_name = file.file_name.lower()
        if not (file_name.endswith('.xlsx') or file_name.endswith('.xls') or file_name.endswith('.csv')):
            self.bot.reply_to(
                message,
                "❌ Please upload only Excel (.xlsx, .xls) or CSV (.csv) files."
            )
            return
        try:
            file_info = self.bot.get_file(file.file_id)
            downloaded_file = self.bot.download_file(file_info.file_path)
            file_data = BytesIO(downloaded_file)
            if file_name.endswith('.csv'):
                df = pd.read_csv(file_data)
            else:
                df = pd.read_excel(file_data)
            village_col = None
            for col in df.columns:
                if 'village' in col.lower():
                    village_col = col
                    break
            if village_col is None:
                self.bot.reply_to(message, "❌ No 'Village' column found in the file.")
                return
            seen = set()
            filtered_villages = []
            skipped = []
            for v in df[village_col].dropna().astype(str).tolist():
                v_clean = v.strip().title()
                if not v_clean or v_clean.startswith('/') or not v_clean.replace(' ', '').isalnum():
                    logger.warning(f"Skipping invalid village name: '{v_clean}'")
                    skipped.append(v_clean)
                    continue
                v_key = v_clean.casefold()
                if v_key in seen:
                    continue
                seen.add(v_key)
                filtered_villages.append(v_clean)
            if not filtered_villages:
                self.bot.reply_to(message, "❌ No valid villages found in the file.")
                return
            logger.info(f"User {user_id} uploaded villages: {filtered_villages}")
            users_collection.update_one(
                {'user_id': user_id},
                {'$set': {'villages': filtered_villages}},
                upsert=True
            )
            reply_msg = (
                f"✅ Successfully added {len(filtered_villages)} villages!\n\n"
                f"**Villages added:** {', '.join(filtered_villages[:5])}"
                + (f" and {len(filtered_villages)-5} more..." if len(filtered_villages) > 5 else "")
            )
            if skipped:
                reply_msg += f"\n\n⚠️ Skipped invalid names: {', '.join(skipped[:5])}" + (f" and {len(skipped)-5} more..." if len(skipped) > 5 else "")
            self.bot.reply_to(message, reply_msg, parse_mode='Markdown')
            # Refresh the settings UI
            user = users_collection.find_one({'user_id': user_id})
            villages = user.get('villages', []) if user else []
            villages_text = (
                '\n'.join([f"{i+1}. {v}" for i, v in enumerate(villages)])
                if villages else 'No villages added yet.'
            )
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(
                types.InlineKeyboardButton("➕ Add Village", callback_data="settings_add_village")
            )
            if villages:
                keyboard.add(
                    types.InlineKeyboardButton("🗑️ Remove Village", callback_data="settings_remove_village")
                )
            keyboard.add(
                types.InlineKeyboardButton("📁 Upload File (Replace All)", callback_data="settings_upload_villages")
            )
            keyboard.add(
                types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
            )
            self.bot.send_message(
                message.chat.id,
                f"🏘️ **Your Villages:**\n\n{villages_text}\n\nYou can add, remove, or upload a new list to replace all.",
                reply_markup=keyboard,
                parse_mode='Markdown'
            )
        except Exception as e:
            logger.error(f"Error processing file for user {user_id}: {e}")
            self.bot.reply_to(message, f"❌ Error processing file: {str(e)}")

    def record_activity_command(self, message):
        """Handle /act command"""
        user_id = message.from_user.id
        user = users_collection.find_one({'user_id': user_id})

        logger.info(f"User {user_id} initiated /act command")

        if not user or not user.get('villages'):
            logger.warning(f"User {user_id} has no villages configured")
            self.bot.reply_to(
                message, "❌ Please add villages first using /settings command."
            )
            return

        args = message.text.split()[1:] if len(message.text.split()) > 1 else []
        date_str = datetime.now(IST).strftime('%d/%m/%Y')
        purpose = None

        if args:
            try:
                datetime.strptime(args[0], '%d/%m/%Y')
                date_str = args[0]
                purpose = ' '.join(args[1:]) if len(args) > 1 else None
                logger.info(f"Date parsed from args: {date_str}, Purpose: {purpose}")
            except ValueError:
                purpose = ' '.join(args)
                logger.info(f"Purpose from args: {purpose}")

        logger.info(f"Showing village buttons to user {user_id}")
        self.show_village_buttons(message, user['villages'])

    def show_village_buttons(self, message, villages: List[str]):
        """Show village selection buttons with filtering and additional options"""
        user_id = message.from_user.id
        user = users_collection.find_one({'user_id': user_id})

        # Ensure all village names are in proper case for display and comparison
        villages = [v.title() for v in villages]

        current_time = datetime.now(IST)
        current_month = current_time.month
        current_year = current_time.year
        covered_villages = set()

        # Migrate to new structure if needed
        user_id = message.from_user.id
        self.migrate_activities_structure(user_id)

        # Get activities from new structure
        activities = user.get('activities', {})
        year_str = str(current_year)
        month_str = str(current_month)

        if year_str in activities and month_str in activities[year_str]:
            for activity in activities[year_str][month_str]:
                try:
                    activity_date = datetime.strptime(activity['date'], '%d/%m/%Y')
                    if activity_date.month == current_month and activity_date.year == current_year:
                        if activity.get('to_village'):
                            covered_villages.add(activity['to_village'].title())
                except ValueError:
                    continue

        available_villages = [v for v in villages if v not in covered_villages]

        total_villages = len(villages)

        covered_text = ""
        # Filter out empty village names which might be present for holidays/HQ days
        valid_covered_villages = sorted([v for v in covered_villages if v])
        if valid_covered_villages:
            covered_villages_str = ", ".join(valid_covered_villages)
            covered_text = (
                f"\n\n✅ Already covered this month: {len(valid_covered_villages)} villages out of {total_villages}."
                f"\nVisited: {covered_villages_str}"
            )

        keyboard = types.InlineKeyboardMarkup()
        headquarters = user.get('headquarters', 'HQ')
        keyboard.add(
            types.InlineKeyboardButton(
                f"🏢 {headquarters} (headquarters)", callback_data=f"village_{headquarters}"
            )
        )

        for i in range(0, len(available_villages), 2):
            row = []
            row.append(
                types.InlineKeyboardButton(
                    available_villages[i], callback_data=f"village_{available_villages[i]}"
                )
            )
            if i + 1 < len(available_villages):
                row.append(
                    types.InlineKeyboardButton(
                        available_villages[i + 1], callback_data=f"village_{available_villages[i + 1]}"
                    )
                )
            keyboard.add(*row)

        keyboard.add(
            types.InlineKeyboardButton("✏️ Manual Entry", callback_data="village_manual")
        )
        keyboard.add(
            types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
        )

        today_date = current_time.strftime('%d/%m/%Y')
        sent = self.bot.send_message(
            message.chat.id,
            f"🏘️ Select the village you visited today {today_date}:{covered_text}",
            reply_markup=keyboard
        )
        # Track pending prompt for timeout
        self.pending_prompts[message.from_user.id] = {
            'message_id': sent.message_id,
            'chat_id': message.chat.id,
            'timeout_time': time.time() + USER_TIMEOUT,
            'type': 'button',
        }

    def handle_village_selection(self, message, temp_activity: Dict, timeout=USER_TIMEOUT):
        """Handle village selection (text input for next step)"""
        logger.info(f"Village selection handler called with text: {message.text}")
        village = message.text.strip().title()
        temp_activity['to_village'] = village

        logger.info(f"Village selected: {village}, temp_activity: {temp_activity}")

        if temp_activity.get('purpose'):
            logger.info("Purpose already exists, saving activity")
            self.save_activity(message, temp_activity)
        else:
            logger.info("No purpose, showing purpose buttons and registering next step handler for custom purpose only")
            # Only register next step handler for custom purpose, not for all
            def custom_purpose_handler(msg):
                # This handler is only called if user selects custom purpose and enters text
                self.handle_purpose_selection(msg, temp_activity, timeout=timeout)

            # Get the month from the activity date if it exists, otherwise use current month
            if 'date' in temp_activity:
                try:
                    activity_date = datetime.strptime(temp_activity['date'], '%d/%m/%Y')
                    month = activity_date.month  # Get month as integer
                    logger.debug(f"Extracted month {month} from activity date: {temp_activity['date']}")
                except ValueError as e:
                    logger.error(f"Error parsing date {temp_activity['date']}: {e}")
                    month = datetime.now(IST).month
                    logger.debug(f"Using current month {month} due to date parsing error")
            else:
                month = datetime.now(IST).month
                logger.debug(f"Using current month: {month}")

            logger.debug(f"MAIN_ACTIVITIES_BY_MONTH[{month}]: {MAIN_ACTIVITIES_BY_MONTH.get(month, [])}")

            # Show purpose buttons with the correct month's default purposes
            self.show_purpose_buttons_with_custom_handler(message, temp_activity, custom_purpose_handler, month=month)

        # Remove pending prompt on user response
        self.pending_prompts.pop(message.from_user.id, None)

    def show_purpose_buttons_with_custom_handler(self, message, temp_activity, custom_purpose_handler, month=None):
        """Show purpose selection buttons with numbered activities, and only register next step handler for custom purpose"""
        user_id = message.from_user.id if hasattr(message, 'from_user') else message.chat.id

        # Get month from temp_activity date if available, otherwise use provided month or current month
        if 'date' in temp_activity:
            activity_date = datetime.strptime(temp_activity['date'], '%d/%m/%Y')
            month = int(activity_date.month)
            logger.debug(f"Using month {month} from temp_activity date: {temp_activity['date']}")
        elif month is not None:
            month = int(month)
            logger.debug(f"Using provided month: {month}")
        else:
            month = int(datetime.now(IST).month)
            logger.debug(f"Using current month: {month}")

        logger.debug(f"Month value: {month} (type: {type(month)})")

        # Get current date for display
        current_date = temp_activity.get('date', datetime.now(IST).strftime('%d/%m/%Y'))
        logger.debug(f"Using date for display: {current_date}")

        # Get all activities for the user and month
        user_activities = self.get_user_activities(user_id, month)
        logger.debug(f"Retrieved activities for user {user_id} and month {month}: {user_activities}")

        # Log activity information
        logger.info(f"User {user_id}: Using activities for date {current_date} (month: {calendar.month_name[month]})")
        logger.info(f"Activities for {calendar.month_name[month]}: {user_activities}")

        keyboard = types.InlineKeyboardMarkup()

        # Show numbered activities with numbered buttons in rows (side by side)
        row = []
        for i, purpose in enumerate(user_activities, 1):
            row.append(types.InlineKeyboardButton(f"{i}", callback_data=f"purpose_idx_{i-1}"))
            # Add 5 buttons per row
            if len(row) == 5:
                keyboard.add(*row)
                row = []

        # Add remaining buttons if any
        if row:
            keyboard.add(*row)

        keyboard.add(
            types.InlineKeyboardButton("📝 Manual Entry", callback_data="purpose_custom")
        )
        keyboard.add(
            types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
        )

        # Create numbered list text
        activities_text = '\n'.join([f"{i}. {purpose}" for i, purpose in enumerate(user_activities, 1)])

        message_text = f"🎯 **Select the purpose of visit for {current_date}**\n\n**Activities for {calendar.month_name[month]}:**\n{activities_text}\n\nClick the number button or use Manual Entry."

        sent = self.bot.send_message(
            message.chat.id,
            message_text,
            reply_markup=keyboard,
            parse_mode='Markdown'
        )
        self.pending_prompts[message.from_user.id] = {
            'message_id': sent.message_id,
            'chat_id': message.chat.id,
            'timeout_time': time.time() + USER_TIMEOUT,
            'type': 'button',
        }

        # Register a callback handler for the purpose buttons
        @self.bot.callback_query_handler(func=lambda call: call.from_user.id == user_id and call.data.startswith('purpose_'))
        def purpose_callback(call):
            if call.data == 'purpose_custom':
                # Register next step handler for custom purpose text input
                self.bot.edit_message_text(
                    "✏️ Please type your custom purpose:",
                    call.message.chat.id,
                    call.message.message_id
                )
                self.bot.register_next_step_handler(
                    call.message,
                    custom_purpose_handler
                )
            elif call.data.startswith('purpose_idx_'):
                # Handle numbered activity selection
                try:
                    idx = int(call.data.replace('purpose_idx_', ''))
                    if 0 <= idx < len(user_activities):
                        purpose = user_activities[idx]
                        logger.info(f"Purpose selected by index {idx}: {purpose}")
                        temp_activity['purpose'] = purpose
                        temp_activity['user_id'] = user_id

                        # Delete the purpose selection message before saving
                        try:
                            self.bot.delete_message(call.message.chat.id, call.message.message_id)
                            logger.debug(f"Deleted purpose selection message for user {user_id}")
                        except Exception as e:
                            logger.error(f"Failed to delete purpose selection message for user {user_id}: {e}")

                        self.save_activity_callback(call, temp_activity)
                    else:
                        logger.error(f"Invalid purpose index {idx} for user {user_id}")
                        self.bot.answer_callback_query(call.id, "❌ Invalid selection. Please try again.")
                        return
                except ValueError:
                    logger.error(f"Invalid purpose index format: {call.data}")
                    self.bot.answer_callback_query(call.id, "❌ Invalid selection. Please try again.")
                    return
            else:
                # Handle legacy purpose selection (fallback)
                purpose = call.data.replace('purpose_', '')
                logger.info(f"Purpose selected (legacy): {purpose}")
                temp_activity['purpose'] = purpose
                temp_activity['user_id'] = user_id

                # Delete the purpose selection message before saving
                try:
                    self.bot.delete_message(call.message.chat.id, call.message.message_id)
                    logger.debug(f"Deleted purpose selection message for user {user_id}")
                except Exception as e:
                    logger.error(f"Failed to delete purpose selection message for user {user_id}: {e}")

                self.save_activity_callback(call, temp_activity)

            if user_id in self.callback_data:
                del self.callback_data[user_id]
                logger.info(f"Cleaned up callback_data for user {user_id}")
                return

    def show_purpose_buttons(self, message, user_id=None, month=None):
        """Show purpose selection buttons with numbered activities"""
        if user_id is None:
            user_id = message.from_user.id if hasattr(message, 'from_user') else message.chat.id
        user_activities = self.get_user_activities(user_id, month)

        logger.info(f"show_purpose_buttons called for user {user_id}, got {len(user_activities)} activities: {user_activities}")

        keyboard = types.InlineKeyboardMarkup()

        # Show numbered activities with numbered buttons in rows (side by side)
        row = []
        for i, purpose in enumerate(user_activities, 1):
            row.append(types.InlineKeyboardButton(f"{i}", callback_data=f"purpose_idx_{i-1}"))
            # Add 5 buttons per row
            if len(row) == 5:
                keyboard.add(*row)
                row = []

        # Add remaining buttons if any
        if row:
            keyboard.add(*row)

        keyboard.add(
            types.InlineKeyboardButton("📝 Manual Entry", callback_data="purpose_custom")
        )
        keyboard.add(
            types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
        )

        # Create numbered list text
        activities_text = '\n'.join([f"{i}. {purpose}" for i, purpose in enumerate(user_activities, 1)])
        message_text = f"🎯 **Select the purpose of visit:**\n\n{activities_text}\n\nClick the number button or use Manual Entry."

        sent = self.bot.send_message(
            message.chat.id,
            message_text,
            reply_markup=keyboard,
            parse_mode='Markdown'
        )
        # Track pending prompt for timeout
        self.pending_prompts[message.from_user.id] = {
            'message_id': sent.message_id,
            'chat_id': message.chat.id,
            'timeout_time': time.time() + USER_TIMEOUT,
            'type': 'button',
        }

    def show_purpose_buttons_edit(self, message):
        """Edit message to show purpose selection buttons with numbered activities"""
        user_id = message.chat.id
        user_activities = self.get_user_activities(user_id)

        keyboard = types.InlineKeyboardMarkup()

        # Show numbered activities with numbered buttons in rows (side by side)
        row = []
        for i, purpose in enumerate(user_activities, 1):
            row.append(types.InlineKeyboardButton(f"{i}", callback_data=f"purpose_idx_{i-1}"))
            # Add 5 buttons per row
            if len(row) == 5:
                keyboard.add(*row)
                row = []

        # Add remaining buttons if any
        if row:
            keyboard.add(*row)

        keyboard.add(
            types.InlineKeyboardButton("📝 Manual Entry", callback_data="purpose_custom")
        )
        keyboard.add(
            types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
        )

        # Create numbered list text
        activities_text = '\n'.join([f"{i}. {purpose}" for i, purpose in enumerate(user_activities, 1)])
        message_text = f"🎯 **Select the purpose of visit:**\n\n{activities_text}\n\nClick the number button or use Manual Entry."

        try:
            self.bot.edit_message_text(
                message_text,
                message.chat.id,
                message.message_id,
                reply_markup=keyboard,
                parse_mode='Markdown'
            )
            # Track pending prompt for timeout
            self.pending_prompts[message.chat.id] = {
                'message_id': message.message_id,
                'chat_id': message.chat.id,
                'timeout_time': time.time() + USER_TIMEOUT,
                'type': 'button',
            }
        except Exception as e:
            logger.error(f"Error editing message for user {user_id}: {e}")
            sent = self.bot.send_message(
                message.chat.id,
                message_text,
                reply_markup=keyboard,
                parse_mode='Markdown'
            )
            self.pending_prompts[message.chat.id] = {
                'message_id': sent.message_id,
                'chat_id': message.chat.id,
                'timeout_time': time.time() + USER_TIMEOUT,
                'type': 'button',
            }

    def handle_purpose_selection(self, message, temp_activity: Dict, timeout=USER_TIMEOUT):
        """Handle purpose selection (text input for custom purpose)"""
        logger.info(f"Purpose selection handler called with text: {message.text}")
        logger.info(f"Current temp_activity: {temp_activity}")
        purpose = message.text.strip()
        temp_activity['purpose'] = purpose
        logger.info(f"Purpose set to: {purpose}")
        self.save_activity(message, temp_activity)

        # Remove pending prompt on user response
        self.pending_prompts.pop(message.from_user.id, None)

    def save_activity(self, message, activity_data: Dict):
        """Save activity to database"""
        user_id = message.from_user.id
        user = users_collection.find_one({'user_id': user_id})

        logger.info(f"Saving activity for user {user_id}: {activity_data}")

        if not user:
            logger.error(f"User {user_id} not found in database")
            return

        if 'date' not in activity_data:
            activity_data['date'] = datetime.now(IST).strftime('%d/%m/%Y')
            logger.info(f"Added default date: {activity_data['date']}")

        if 'to_village' not in activity_data:
            logger.error(f"Village information missing for user {user_id}")
            self.bot.reply_to(
                message, "❌ Village information is missing. Please try again."
            )
            return

        if 'purpose' not in activity_data:
            logger.error(f"Purpose information missing for user {user_id}")
            self.bot.reply_to(
                message, "❌ Purpose information is missing. Please try again."
            )
            return

        activity = {
            'date': activity_data['date'],
            'from': user['headquarters'] or 'HQ',
            'to_village': activity_data['to_village'],
            'purpose': activity_data['purpose']
        }

        logger.info(f"Created activity object: {activity}")

        # Migrate to new structure if needed
        self.migrate_activities_structure(user_id)

        # Parse date to get year and month
        try:
            dt = datetime.strptime(activity['date'], '%d/%m/%Y')
            year = str(dt.year)
            month = str(dt.month)
        except Exception as e:
            logger.error(f"Invalid date format for activity: {activity['date']}")
            self.bot.reply_to(message, "❌ Invalid date format. Please try again.")
            return

        # Get current activities structure
        user = users_collection.find_one({'user_id': user_id})
        activities = user.get('activities', {})

        # Check if activity for this date already exists
        existing_activity = None
        if year in activities and month in activities[year]:
            for i, act in enumerate(activities[year][month]):
                if act['date'] == activity['date']:
                    existing_activity = i
                    break

        if existing_activity is not None:
            # Update existing activity
            users_collection.update_one(
                {'user_id': user_id},
                {'$set': {f'activities.{year}.{month}.{existing_activity}': activity}}
            )
            message_text = "✅ Activity updated successfully!"
        else:
            # Add new activity
            users_collection.update_one(
                {'user_id': user_id},
                {'$push': {f'activities.{year}.{month}': activity}}
            )
            message_text = "✅ Activity recorded successfully!"

        activity_summary = (
            f"**Date:** {activity['date']}\n"
            f"**From:** {activity['from']}\n"
            f"**To:** {activity['to_village']}\n"
            f"**Purpose:** {activity['purpose']}"
        )

        # Delete the daily prompt message if it exists
        if user_id in self.daily_prompt_message_ids:
            try:
                msg_info = self.daily_prompt_message_ids[user_id]
                if isinstance(msg_info, dict):
                    message_id = msg_info['message_id']
                else:
                    # Handle legacy format where only message_id was stored
                    message_id = msg_info
                self.bot.delete_message(message.chat.id, message_id)
                logger.info(f"🗑️ Deleted daily prompt message for user {user_id}")
                del self.daily_prompt_message_ids[user_id]
            except Exception as e:
                logger.error(f"Error deleting daily prompt message for user {user_id}: {e}")

        self.bot.send_message(
            message.chat.id,
            f"{message_text}\n\n{activity_summary}",
            parse_mode='Markdown'
        )

    def save_activity_callback(self, call, temp_activity):
        """Save activity to database from callback query"""
        user_id = temp_activity['user_id']
        logger.info(f"Saving activity from callback for user {user_id}: {temp_activity}")
        user = users_collection.find_one({'user_id': user_id})

        if not user:
            logger.error(f"User {user_id} not found in database")
            return

        if 'date' not in temp_activity:
            temp_activity['date'] = datetime.now(IST).strftime('%d/%m/%Y')
            logger.info(f"Added default date: {temp_activity['date']}")

        if 'to_village' not in temp_activity:
            logger.error(f"Village information missing for user {user_id}")
            self.bot.send_message(
                call.message.chat.id,
                "❌ Village information is missing. Please try again."
            )
            return

        if 'purpose' not in temp_activity:
            logger.error(f"Purpose information missing for user {user_id}")
            self.bot.send_message(
                call.message.chat.id,
                "❌ Purpose information is missing. Please try again."
            )
            return

        activity = {
            'date': temp_activity['date'],
            'from': user['headquarters'] or 'HQ',
            'to_village': temp_activity['to_village'],
            'purpose': temp_activity['purpose']
        }

        logger.info(f"Created activity object: {activity}")

        # Migrate to new structure if needed
        self.migrate_activities_structure(user_id)

        # Parse date to get year and month
        try:
            dt = datetime.strptime(activity['date'], '%d/%m/%Y')
            year = str(dt.year)
            month = str(dt.month)
        except Exception as e:
            logger.error(f"Invalid date format for activity: {activity['date']}")
            self.bot.send_message(call.message.chat.id, "❌ Invalid date format. Please try again.")
            return

        # Get current activities structure
        user = users_collection.find_one({'user_id': user_id})
        activities = user.get('activities', {})

        # Check if activity for this date already exists
        existing_activity = None
        if year in activities and month in activities[year]:
            for i, act in enumerate(activities[year][month]):
                if act['date'] == activity['date']:
                    existing_activity = i
                    break

        if existing_activity is not None:
            # Update existing activity
            users_collection.update_one(
                {'user_id': user_id},
                {'$set': {f'activities.{year}.{month}.{existing_activity}': activity}}
            )
            message_text = "✅ Activity updated successfully!"
        else:
            # Add new activity
            users_collection.update_one(
                {'user_id': user_id},
                {'$push': {f'activities.{year}.{month}': activity}}
            )
            message_text = "✅ Activity recorded successfully!"

        activity_summary = (
            f"**Date:** {activity['date']}\n"
            f"**From:** {activity['from']}\n"
            f"**To:** {activity['to_village']}\n"
            f"**Purpose:** {activity['purpose']}"
        )

        # Delete the purpose selection prompt before sending confirmation
        try:
            self.bot.delete_message(call.message.chat.id, call.message.message_id)
            logger.debug(f"Deleted purpose selection prompt for user {user_id}")
        except Exception as e:
            logger.error(f"Failed to delete purpose selection prompt for user {user_id}: {e}")

        # Delete the daily prompt message if it exists
        if user_id in self.daily_prompt_message_ids:
            try:
                msg_info = self.daily_prompt_message_ids[user_id]
                if isinstance(msg_info, dict):
                    message_id = msg_info['message_id']
                else:
                    # Handle legacy format where only message_id was stored
                    message_id = msg_info
                self.bot.delete_message(call.message.chat.id, message_id)
                logger.info(f"🗑️ Deleted daily prompt message for user {user_id}")
                del self.daily_prompt_message_ids[user_id]
            except Exception as e:
                logger.error(f"Error deleting daily prompt message for user {user_id}: {e}")

        # Send confirmation message
        self.bot.send_message(
            call.message.chat.id,
            f"{message_text}\n\n{activity_summary}",
            parse_mode='Markdown'
        )

        # Remove pending prompt on user response
        self.pending_prompts.pop(user_id, None)

    def download_activities(self, message):
        """Handle /dnact command"""
        user_id = message.from_user.id
        user = users_collection.find_one({'user_id': user_id})

        if not user or not user.get('activities'):
            self.bot.reply_to(message, "❌ No activities found.")
            return

        args = message.text.split()[1:] if len(message.text.split()) > 1 else []
        if len(args) < 2:
            self.bot.reply_to(
                message,
                "❌ Please specify month and year. Usage: /dnact <month_number> <year> (e.g., /dnact 6 2024)"
            )
            return
        try:
            month_filter = int(args[0])
            year_filter = int(args[1])
            if month_filter < 1 or month_filter > 12:
                raise ValueError
        except ValueError:
            self.bot.reply_to(
                message,
                "❌ Invalid month or year. Usage: /dnact <month_number> <year> (e.g., /dnact 6 2024)"
            )
            return

        # Migrate to new structure if needed
        self.migrate_activities_structure(user_id)

        # Get activities from new nested structure
        activities = user.get('activities', {})
        year_str = str(year_filter)
        month_str = str(month_filter)

        if year_str not in activities or month_str not in activities[year_str]:
            self.bot.reply_to(
                message,
                f"❌ No activities found for month {month_filter} and year {year_filter}."
            )
            return

        activities = activities[year_str][month_str]

        if not activities:
            self.bot.reply_to(
                message,
                f"❌ No activities found for month {month_filter} and year {year_filter}."
            )
            return

        # Only count as tour days if to_village is not empty
        tour_days = sum(1 for act in activities if act.get('to_village', '').strip())
        min_required = 20

        month_name = calendar.month_name[month_filter]
        status_msg = f"📊 **{month_name} {year_filter} Tour Summary**\n"
        status_msg += f"Tour Days: {tour_days}\n"
        status_msg += f"Required: {min_required}\n"
        status_msg += (
            f"⚠️ Short by {min_required - tour_days} days"
            if tour_days < min_required
            else "✅ Requirement met"
        )
        self.bot.reply_to(message, status_msg, parse_mode='Markdown')

        df = pd.DataFrame(activities)
        csv_buffer = BytesIO()
        df.to_csv(csv_buffer, index=False)
        csv_buffer.seek(0)

        filename = (
            f"tour_activities_{month_filter}_{year_filter}_"
            f"{datetime.now(IST).strftime('%Y%m%d')}.csv"
        )
        self.bot.send_document(
            message.chat.id,
            document=(filename, csv_buffer),
            caption=f"📋 Tour Activities Report\n{len(activities)} activities exported"
        )

    def settings_command(self, message):
        """Handle /settings command"""
        user_id = message.from_user.id
        user = users_collection.find_one({'user_id': user_id})

        if not user:
            users_collection.insert_one({
                'user_id': user_id,
                'headquarters': None,
                'villages': [],
                'activities': [],
                'custom_activities': [],
                'role': None
            })
            user = users_collection.find_one({'user_id': user_id})

        hq_status = (
            f"✅ {user.get('headquarters', 'Not set')}"
            if user.get('headquarters')
            else "❌ Not set"
        )
        role = user.get('role')
        role_status = f"✅ {role}" if role else "❌ Not set"
        villages_count = len(user.get('villages', []))
        custom_activities_count = len(user.get('custom_activities', []))
        default_purpose = user.get('default_purpose') or 'Not set'
        holidays_count = len(user.get('public_holidays', []))

        keyboard_buttons = [
            types.InlineKeyboardButton("👤 Set Role", callback_data="settings_setrole"),
            types.InlineKeyboardButton("🏢 Set Headquarters", callback_data="settings_sethq"),
            types.InlineKeyboardButton("🏘️ Add Villages", callback_data="settings_addvil"),
            types.InlineKeyboardButton("📋 Manage Activities", callback_data="settings_activities"),
            types.InlineKeyboardButton("🎯 Default Purpose", callback_data="settings_default_purpose"),
            types.InlineKeyboardButton("📅 Add Public Holidays", callback_data="settings_upload_holidays"),
            types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
        ]
        keyboard = types.InlineKeyboardMarkup()
        for btn in keyboard_buttons:
            keyboard.add(btn)

        settings_text = (
            f"⚙️ **Settings**\n\n"
            f"**Headquarters:** {hq_status}\n"
            f"**Villages:** {villages_count} added\n"
            f"**Custom Activities:** {custom_activities_count} defined\n"
            f"**Default Purpose:** {default_purpose}\n"
            f"**Public Holidays:** {holidays_count} added\n\n"
            f"Select an option to configure:"
        )

        try:
            self.bot.send_message(
                message.chat.id,
                settings_text,
                reply_markup=keyboard,
                parse_mode='Markdown'
            )
        except Exception as e:
            logger.error(f"Error sending settings menu to user {user_id}: {e}")
            self.bot.send_message(
                message.chat.id,
                "❌ Error displaying settings. Please try again."
            )

    def owner_settings_command(self, message):
        """Handle /ownerset command for bot owner."""
        user_id = message.from_user.id
        if str(user_id) != str(OWNER_ID):
            self.bot.reply_to(message, "❌ You are not authorized to use this command.")
            return

        settings_text = (
            f"👑 **Owner Settings**\n\n"
            f"Here you can configure bot-wide settings.\n\n"
            f"**Current Schedule Times (IST):**\n"
            f"• Daily Prompt: `{DAILY_PROMPT_TIME}`\n"
            f"• Default Activity Fallback: `{DEFAULT_ACTIVITY_TIME}`"
        )

        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(
            types.InlineKeyboardButton("⏰ Change Daily Prompt Time", callback_data="owner_set_prompt_time")
        )
        keyboard.add(
            types.InlineKeyboardButton("🤖 Change Default Activity Time", callback_data="owner_set_fallback_time")
        )
        keyboard.add(
            types.InlineKeyboardButton("❌ Close", callback_data="cancel_selection")
        )

        self.bot.send_message(
            message.chat.id,
            settings_text,
            reply_markup=keyboard,
            parse_mode='Markdown'
        )

    def _is_valid_time_format(self, time_str):
        """Validate HH:MM format."""
        try:
            datetime.strptime(time_str, '%H:%M')
            return True
        except ValueError:
            return False

    def handle_owner_set_prompt_time(self, message, timeout=None):
        """Handle owner input for new daily prompt time."""
        user_id = message.from_user.id
        if str(user_id) != str(OWNER_ID):
            return

        new_time = message.text.strip()
        if not self._is_valid_time_format(new_time):
            self.bot.reply_to(message, "❌ Invalid format. Please use HH:MM (e.g., 19:30).")
            logger.warning(f"Invalid time format received from user {user_id}: {new_time}")
            self.bot.register_next_step_handler(message, self.handle_owner_set_prompt_time)
            return

        try:
            config_collection.update_one(
                {'_id': 'schedule_times'},
                {'$set': {'daily_prompt_time': new_time}},
                upsert=True
            )
            global DAILY_PROMPT_TIME
            DAILY_PROMPT_TIME = new_time
            
            # Delete the original prompt message if present
            if user_id in self.input_prompt_message:
                try:
                    self.bot.delete_message(message.chat.id, self.input_prompt_message[user_id])
                except Exception as e:
                    logger.error(f"Error deleting prompt message for user {user_id}: {e}")
                del self.input_prompt_message[user_id]
            
            # Delete the user's input message
            try:
                self.bot.delete_message(message.chat.id, message.message_id)
            except Exception as e:
                logger.error(f"Error deleting user input message for user {user_id}: {e}")
                
            self.bot.send_message(message.chat.id, f"✅ Daily prompt time updated to **{new_time}** IST. The change will take effect on the next schedule check.", parse_mode='Markdown')
            logger.info(f"Owner updated daily prompt time to {new_time}")
        except Exception as e:
            logger.error(f"Error updating prompt time in DB: {e}")
            self.bot.reply_to(message, "❌ An error occurred while saving the new time.")

    def handle_owner_set_fallback_time(self, message, timeout=None):
        """Handle owner input for new default activity fallback time."""
        user_id = message.from_user.id
        if str(user_id) != str(OWNER_ID):
            return

        new_time = message.text.strip()
        if not self._is_valid_time_format(new_time):
            self.bot.reply_to(message, "❌ Invalid format. Please use HH:MM (e.g., 20:00).")
            self.bot.register_next_step_handler(message, self.handle_owner_set_fallback_time)
            return

        try:
            config_collection.update_one(
                {'_id': 'schedule_times'},
                {'$set': {'default_activity_time': new_time}},
                upsert=True
            )
            global DEFAULT_ACTIVITY_TIME
            DEFAULT_ACTIVITY_TIME = new_time
            
            # Delete the original prompt message if present
            if user_id in self.input_prompt_message:
                try:
                    self.bot.delete_message(message.chat.id, self.input_prompt_message[user_id])
                except Exception as e:
                    logger.error(f"Error deleting prompt message for user {user_id}: {e}")
                del self.input_prompt_message[user_id]
            
            # Delete the user's input message
            try:
                self.bot.delete_message(message.chat.id, message.message_id)
            except Exception as e:
                logger.error(f"Error deleting user input message for user {user_id}: {e}")
                
            self.bot.send_message(message.chat.id, f"✅ Default activity fallback time updated to **{new_time}** IST. The change will take effect on the next schedule check.", parse_mode='Markdown')
            logger.info(f"Owner updated default activity time to {new_time}")
        except Exception as e:
            logger.error(f"Error updating fallback time in DB: {e}")
            self.bot.reply_to(message, "❌ An error occurred while saving the new time.")

    def handle_settings_sethq(self, message, timeout=USER_TIMEOUT):
        """Handle headquarters setting from settings"""
        user_id = message.from_user.id
        if user_id in self.cancelled_users:
            self.cancelled_users.remove(user_id)
            return
        headquarters = message.text.strip().title()

        if not headquarters:
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(
                types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
            )
            sent = self.bot.send_message(
                message.chat.id,
                f"❌ Please provide a valid headquarters name.\n\n⚠️ You have {timeout} seconds to reply before the request times out.",
                reply_markup=keyboard
            )
            logger.info(f"Prompted user {user_id} for headquarters input, message_id: {sent.message_id}")
            self.input_prompt_message[user_id] = sent.message_id
            self.bot.register_next_step_handler(
                message,
                self.handle_settings_sethq,
                timeout=timeout
            )
            return

        try:
            users_collection.update_one(
                {'user_id': user_id},
                {'$set': {'headquarters': headquarters}},
                upsert=True
            )
        except Exception as e:
            logger.error(f"Error updating headquarters for user {user_id}: {e}")
            self.bot.send_message(
                message.chat.id,
                "❌ Error saving headquarters. Please try again."
            )
            return

        # Delete the original prompt message if present
        if user_id in self.input_prompt_message:
            try:
                self.bot.delete_message(message.chat.id, self.input_prompt_message[user_id])
            except Exception as e:
                logger.error(f"Error deleting prompt message for user {user_id}: {e}")
            del self.input_prompt_message[user_id]
        self.bot.send_message(
            message.chat.id,
            f"✅ Headquarters set to: **{headquarters}**",
            parse_mode='Markdown'
        )

        # Remove pending prompt on user response
        self.pending_prompts.pop(user_id, None)

    def handle_settings_setrole(self, message, timeout=USER_TIMEOUT):
        """Handle role setting from settings"""
        user_id = message.from_user.id
        if user_id in self.cancelled_users:
            self.cancelled_users.remove(user_id)
            return
        role = message.text.strip()  # Preserve case

        if not role:
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(
                types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
            )
            sent = self.bot.send_message(
                message.chat.id,
                f"❌ Please provide a valid role/designation.\n\n⚠️ You have {timeout} seconds to reply before the request times out.",
                reply_markup=keyboard
            )
            logger.info(f"Prompted user {user_id} for role input, message_id: {sent.message_id}")
            self.input_prompt_message[user_id] = sent.message_id
            self.bot.register_next_step_handler(
                message,
                self.handle_settings_setrole,
                timeout=timeout
            )
            return

        try:
            users_collection.update_one(
                {'user_id': user_id},
                {'$set': {'role': role}},
                upsert=True
            )
        except Exception as e:
            logger.error(f"Error updating role for user {user_id}: {e}")
            self.bot.send_message(
                message.chat.id,
                "❌ Error saving role. Please try again."
            )
            return

        # Delete the original prompt message if present
        if user_id in self.input_prompt_message:
            try:
                self.bot.delete_message(message.chat.id, self.input_prompt_message[user_id])
            except Exception as e:
                logger.error(f"Error deleting prompt message for user {user_id}: {e}")
            del self.input_prompt_message[user_id]
        self.bot.send_message(
            message.chat.id,
            f"✅ Role set to: **{role}**",
            parse_mode='Markdown'
        )

        # Remove pending prompt on user response
        self.pending_prompts.pop(user_id, None)

    def handle_settings_default_purpose(self, message, timeout=USER_TIMEOUT):
        """Handle default purpose setting from settings"""
        user_id = message.from_user.id
        logger.info(f"handle_settings_default_purpose called for user {user_id} with text: {repr(message.text)}")
        if user_id in self.cancelled_users:
            self.cancelled_users.remove(user_id)
            return
        purpose = message.text.strip()

        if not purpose:
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(
                types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
            )
            sent = self.bot.send_message(
                message.chat.id,
                f"❌ Please provide a valid default purpose.\n\n⚠️ You have {timeout} seconds to reply before the request times out.",
                reply_markup=keyboard
            )
            logger.info(f"Prompted user {user_id} for default purpose input, message_id: {sent.message_id}")
            self.input_prompt_message[user_id] = sent.message_id
            self.bot.register_next_step_handler(
                message,
                self.handle_settings_default_purpose,
                timeout=timeout
            )
            return

        try:
            users_collection.update_one(
                {'user_id': user_id},
                {'$set': {'default_purpose': purpose}},
                upsert=True
            )
        except Exception as e:
            logger.error(f"Error updating default purpose for user {user_id}: {e}")
            self.bot.send_message(
                message.chat.id,
                "❌ Error saving default purpose. Please try again."
            )
            return

        # Delete the original prompt message if present
        if user_id in self.input_prompt_message:
            try:
                logger.info(f"Deleting prompt message for user {user_id}: {self.input_prompt_message[user_id]}")
                self.bot.delete_message(message.chat.id, self.input_prompt_message[user_id])
            except Exception as e:
                logger.error(f"Error deleting prompt message for user {user_id}: {e}")
            del self.input_prompt_message[user_id]

        self.bot.send_message(
            message.chat.id,
            f"✅ Default purpose set to: **{purpose}**\n\nThis purpose will be used for {DEFAULT_ACTIVITY_TIME} auto-entries.",
            parse_mode='Markdown'
        )

        # Remove pending prompt on user response
        self.pending_prompts.pop(user_id, None)

    def handle_settings_add_activity(self, message, timeout=USER_TIMEOUT):
        """Handle adding custom activity from settings"""
        user_id = message.from_user.id
        logger.info(f"handle_settings_add_activity called for user {user_id} with text: {repr(message.text)}")
        if user_id in self.cancelled_users:
            self.cancelled_users.remove(user_id)
            return
        activity = message.text.strip()

        if not activity:
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(
                types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
            )
            sent = self.bot.send_message(
                message.chat.id,
                f"❌ Please provide a valid activity name.\n\n⚠️ You have {timeout} seconds to reply before the request times out.",
                reply_markup=keyboard
            )
            logger.info(f"Prompted user {user_id} for custom activity input, message_id: {sent.message_id}")
            self.input_prompt_message[user_id] = sent.message_id
            self.bot.register_next_step_handler(
                message,
                self.handle_settings_add_activity,
                timeout=timeout
            )
            return

        try:
            users_collection.update_one(
                {'user_id': user_id},
                {'$addToSet': {'custom_activities': activity}},
                upsert=True
            )
        except Exception as e:
            logger.error(f"Error adding custom activity for user {user_id}: {e}")
            self.bot.send_message(
                message.chat.id,
                "❌ Error adding activity. Please try again."
            )
            return

        # Delete the original prompt message if present
        if user_id in self.input_prompt_message:
            try:
                self.bot.delete_message(message.chat.id, self.input_prompt_message[user_id])
            except Exception as e:
                logger.error(f"Error deleting prompt message for user {user_id}: {e}")
            del self.input_prompt_message[user_id]
        self.bot.send_message(
            message.chat.id,
            f"✅ Activity added: **{activity}**",
            parse_mode='Markdown'
        )

        # Remove pending prompt on user response
        self.pending_prompts.pop(user_id, None)

    def get_user_activities(self, user_id, month=None):
        """Get custom activities for a specific user (for purpose selection)"""
        logger.debug(f"get_user_activities called with user_id: {user_id}, month: {month} (type: {type(month) if month is not None else None})")

        # Try to find user with the given user_id
        user = users_collection.find_one({'user_id': user_id})

        # If not found, try with integer conversion (in case user_id is stored as int)
        if not user and isinstance(user_id, str):
            try:
                user = users_collection.find_one({'user_id': int(user_id)})
                logger.info(f"Found user with integer user_id: {int(user_id)}")
            except ValueError:
                pass

        # If still not found, try with string conversion (in case user_id is stored as string)
        if not user and isinstance(user_id, int):
            try:
                user = users_collection.find_one({'user_id': str(user_id)})
                logger.info(f"Found user with string user_id: {str(user_id)}")
            except ValueError:
                pass

        if not user:
            logger.warning(f"No user found for user_id: {user_id} (type: {type(user_id)})")
            return []

        # Get custom activities from settings
        custom_activities = user.get('custom_activities', [])
        logger.debug(f"Retrieved custom activities for user {user_id}: {custom_activities}")

        # If month is not provided, use current month
        if month is None:
            month = int(datetime.now(IST).month)
            logger.debug(f"No month provided, using current month: {month} (type: {type(month)})")
        else:
            # Ensure month is an integer
            month = int(month)
            logger.debug(f"Using provided month: {month} (type: {type(month)})")

        # Get main activities for the specified month
        logger.debug(f"MAIN_ACTIVITIES_BY_MONTH dictionary: {MAIN_ACTIVITIES_BY_MONTH}")
        main_activities = MAIN_ACTIVITIES_BY_MONTH.get(month, [])
        if not main_activities:  # If no activities found for the month, use default
            main_activities = ["survey harvest", "seasonal conditions"]
            logger.debug(f"No activities found for month {month}, using default activities")
        logger.debug(f"Retrieved main activities for month {month}: {main_activities}")

        # Combine custom and main activities, with custom activities taking precedence
        activities = list(set(custom_activities + main_activities))
        logger.debug(f"Returning combined activities for user {user_id}: {activities}")
        return activities

    def daily_prompt(self):
        """Send daily activity prompt to all users"""
        current_time = datetime.now(IST)
        current_date = current_time.strftime('%Y-%m-%d')
        logger.info(f"🕰️ Daily prompt triggered at {current_time.strftime('%H:%M:%S IST')} (scheduled time: {DAILY_PROMPT_TIME})")

        if current_time.weekday() == 6:
            logger.info("📅 Sunday detected - skipping daily prompt (public holiday)")
            return
        if current_time.weekday() == 5:
            first_day = current_time.replace(day=1)
            first_saturday = first_day + timedelta(days=(5 - first_day.weekday()) % 7)
            second_saturday = first_saturday + timedelta(days=7)
            if current_time.date() == second_saturday.date():
                logger.info("📅 Second Saturday detected - skipping daily prompt (public holiday)")
                return

        # Get users with villages configured
        users = users_collection.find({'villages': {'$exists': True, '$ne': []}})  #Recreate the cursor
        user_count = users_collection.count_documents({'villages': {'$exists': True, '$ne': []}})
        logger.info(f"🔍 Found {user_count} users with villages configured")
        for user in users:
            try:
                today_str = current_time.strftime('%d/%m/%Y')
                logger.info(f"👤 Processing user {user['user_id']} for daily prompt")
                
                # Check if daily prompt was already sent today for this user
                user_id = user['user_id']
                if user_id in self.daily_prompt_message_ids:
                    msg_info = self.daily_prompt_message_ids[user_id]
                    if isinstance(msg_info, dict) and msg_info.get('date') == current_date:
                        logger.info(f"Daily prompt already sent today for user {user_id}")
                        continue

                # Migrate to new structure if needed
                # Check if it's a user-defined public holiday today
                is_user_holiday = False
                for h in user.get('public_holidays', []):
                    try:
                        if datetime.strptime(h['date'], '%d/%m/%Y').date() == current_time.date():
                            is_user_holiday = True
                            logger.info(f"📅 User {user['user_id']} - Skipping daily prompt (user holiday: {h['desc']})")
                            break
                    except Exception:
                        continue

                if is_user_holiday:
                    logger.info(f"📅 Skipping daily prompt for user {user['user_id']} because today is a user-defined public holiday.")
                    continue

                self.migrate_activities_structure(user['user_id'])

                # Check if activity exists using new structure
                activities = user.get('activities', {})
                year_str = str(current_time.year)
                month_str = str(current_time.month)

                has_activity = False
                if year_str in activities and month_str in activities[year_str]:
                    has_activity = any(
                        act['date'] == today_str for act in activities[year_str][month_str]
                    )

                if has_activity:
                    logger.info(f"✅ User {user['user_id']} already has activity for today, skipping")
                    continue

                # If user has no default_purpose, select a random activity from MAIN_ACTIVITIES_BY_MONTH
                if not user.get('default_purpose'):
                    logger.info(f"⚠️ User {user['user_id']} has no default_purpose, selecting random activity")
                    current_month = current_time.month
                    logger.debug(f"Current month: {current_month}")

                    # Get activities for current month
                    month_activities = MAIN_ACTIVITIES_BY_MONTH.get(current_month, [])
                    if not month_activities:
                        month_activities = ["survey harvest", "seasonal conditions"]
                    logger.debug(f"Available activities for month {current_month}: {month_activities}")

                    # Select random activity
                    random_purpose = random.choice(month_activities)
                    logger.info(f"Selected random activity for user {user['user_id']}: {random_purpose}")

                    # Save the random purpose as default_purpose temporarily for this activity
                    user['default_purpose'] = random_purpose

                keyboard = types.InlineKeyboardMarkup()
                # Ensure villages are proper case for display and comparison
                villages = [v.title() for v in user.get('villages', [])]

                current_month = current_time.month
                current_year = current_time.year
                covered_villages = set()

                # Get covered villages from new structure
                year_str = str(current_year)
                month_str = str(current_month)
                activities = user.get('activities', {})
                if year_str in activities and month_str in activities[year_str]:
                    for activity in activities[year_str][month_str]:
                        try:
                            activity_date = datetime.strptime(activity['date'], '%d/%m/%Y')
                            if activity_date.month == current_month and activity_date.year == current_year:
                                if activity.get('to_village'):
                                    covered_villages.add(activity['to_village'].title())
                        except ValueError:
                            continue

                available_villages = [v for v in villages if v not in covered_villages]

                headquarters = user.get('headquarters', 'HQ')
                keyboard.add(
                    types.InlineKeyboardButton(
                        f"🏢 {headquarters.title()} (headquarters)", callback_data=f"daily_village_{headquarters.title()}"
                    )
                )

                for i in range(0, len(available_villages), 2):
                    row = []
                    row.append(
                        types.InlineKeyboardButton(
                            available_villages[i], callback_data=f"daily_village_{available_villages[i]}"
                        )
                    )
                    if i + 1 < len(available_villages):
                        row.append(
                            types.InlineKeyboardButton(
                                available_villages[i + 1], callback_data=f"daily_village_{available_villages[i + 1]}"
                            )
                        )
                    keyboard.add(*row)

                keyboard.add(
                    types.InlineKeyboardButton("✏️ Manual Entry", callback_data="daily_village_manual")
                )
                keyboard.add(
                    types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
                )

                # Prepare message based on whether purpose was randomly selected
                purpose_explanation = (
                    "\n\n📝 Note: Since you haven't set a default purpose, I've randomly selected one from this month's activities: **" + user['default_purpose'] + "**"
                    if not user.get('default_purpose') else ""
                )

                logger.info(f"📤 Sending daily prompt to user {user['user_id']} with {len(available_villages)} available villages")
                sent_message = self.bot.send_message(
                    user['user_id'],
                    "🕰️ **Daily Activity Reminder**\n\n"
                    "Please record your tour activity for today. Select the village you visited:" +
                    purpose_explanation,
                    reply_markup=keyboard,
                    parse_mode='Markdown'
                )
                # Store the message ID and date for later deletion and duplicate prevention
                self.daily_prompt_message_ids[user['user_id']] = {
                    'message_id': sent_message.message_id,
                    'date': current_time.strftime('%Y-%m-%d')
                }
            except Exception as e:
                logger.error(f"Error sending daily prompt to user {user['user_id']}: {e}")

        logger.info(f"🏁 Daily prompt function completed")

    def default_activity_fallback(self):
        """Add default activity for users who didn't respond by {DEFAULT_ACTIVITY_TIME}"""
        current_time = datetime.now(IST)
        logger.info(f"🤖 Default activity fallback triggered at {current_time.strftime('%H:%M:%S IST')} (scheduled time: {DEFAULT_ACTIVITY_TIME})")

        if current_time.weekday() == 6:
            # Sunday: public holiday
            today_str = current_time.strftime('%d/%m/%Y')
            users = users_collection.find({'villages': {'$exists': True, '$ne': []}})
            for user in users:
                # Migrate to new structure if needed
                self.migrate_activities_structure(user['user_id'])

                # Check if activity exists using new structure
                activities = user.get('activities', {})
                year_str = str(current_time.year)
                month_str = str(current_time.month)

                has_activity = False
                if year_str in activities and month_str in activities[year_str]:
                    has_activity = any(
                        act['date'] == today_str for act in activities[year_str][month_str]
                    )

                if has_activity:
                    continue

                # Try to get holiday name from user's public_holidays
                holiday_name = None
                for h in user.get('public_holidays', []):
                    try:
                        if datetime.strptime(h['date'], '%d/%m/%Y').date() == current_time.date():
                            holiday_name = h['desc']
                            break
                    except Exception:
                        continue

                if holiday_name:
                    purpose_str = f"Public holiday ({holiday_name})"
                elif current_time.weekday() == 6:
                    purpose_str = "Public holiday (Sunday)"
                elif current_time.weekday() == 5 and current_time.date() == second_saturday.date():
                    purpose_str = "Public holiday (Second Saturday)"
                else:
                    purpose_str = "Public holiday"

                activity = {
                    'date': today_str,
                    'from': user['headquarters'] or 'HQ',
                    'to_village': '',
                    'purpose': purpose_str
                }

                # Save using new structure
                users_collection.update_one(
                    {'user_id': user['user_id']},
                    {'$push': {f'activities.{year_str}.{month_str}': activity}}
                )
                self.bot.send_message(
                    user['user_id'],
                    f"🏖️ **Public Holiday Recorded**\n\n"
                    f"Today is a public holiday (Sunday).\n"
                    f"**Date:** {activity['date']}\n"
                    f"**From:** {activity['from']}\n"
                    f"**To:** {activity['to_village']}\n"
                    f"**Purpose:** {activity['purpose']}\n",
                    parse_mode='Markdown'
                )
            return
        if current_time.weekday() == 5:
            # Check for second Saturday
            first_day = current_time.replace(day=1)
            first_saturday = first_day + timedelta(days=(5 - first_day.weekday()) % 7)
            second_saturday = first_saturday + timedelta(days=7)
            if current_time.date() == second_saturday.date():
                today_str = current_time.strftime('%d/%m/%Y')
                users = users_collection.find({'villages': {'$exists': True, '$ne': []}})
                for user in users:
                    # Migrate to new structure if needed
                    self.migrate_activities_structure(user['user_id'])

                    # Check if activity exists using new structure
                    activities = user.get('activities', {})
                    year_str = str(current_time.year)
                    month_str = str(current_time.month)

                    has_activity = False
                    if year_str in activities and month_str in activities[year_str]:
                        has_activity = any(
                            act['date'] == today_str for act in activities[year_str][month_str]
                        )

                    if has_activity:
                        continue

                    # Try to get holiday name from user's public_holidays
                holiday_name = None
                for h in user.get('public_holidays', []):
                    try:
                        if datetime.strptime(h['date'], '%d/%m/%Y').date() == current_time.date():
                            holiday_name = h['desc']
                            break
                    except Exception:
                        continue

                if holiday_name:
                    purpose_str = f"Public holiday ({holiday_name})"
                elif current_time.weekday() == 6:
                    purpose_str = "Public holiday (Sunday)"
                elif current_time.weekday() == 5 and current_time.date() == second_saturday.date():
                    purpose_str = "Public holiday (Second Saturday)"
                else:
                    purpose_str = "Public holiday"

                activity = {
                    'date': today_str,
                    'from': user['headquarters'] or 'HQ',
                    'to_village': '',
                    'purpose': purpose_str
                }

                # Save using new structure
                users_collection.update_one(
                    {'user_id': user['user_id']},
                    {'$push': {f'activities.{year_str}.{month_str}': activity}}
                )
                self.bot.send_message(
                    user['user_id'],
                    f"🏖️ **Public Holiday Recorded**\n\n"
                    f"Today is a public holiday (Second Saturday).\n"
                    f"**Date:** {activity['date']}\n"
                    f"**From:** {activity['from']}\n"
                    f"**To:** {activity['to_village']}\n"
                    f"**Purpose:** {activity['purpose']}\n",
                    parse_mode='Markdown'
                )
                return

        # Regular weekday default activity
        today_str = current_time.strftime('%d/%m/%Y')
        users = users_collection.find({'villages': {'$exists': True, '$ne': []}})
        for user in users:
            try:
                # Migrate to new structure if needed
                self.migrate_activities_structure(user['user_id'])

                # Check if activity exists using new structure
                activities = user.get('activities', {})
                year_str = str(current_time.year)
                month_str = str(current_time.month)

                has_activity = False
                if year_str in activities and month_str in activities[year_str]:
                    has_activity = any(
                        act['date'] == today_str for act in activities[year_str][month_str]
                    )

                if has_activity:
                    continue

                # If user has no default_purpose, select a random activity from MAIN_ACTIVITIES_BY_MONTH
                if not user.get('default_purpose'):
                    logger.info(f"User {user['user_id']} has no default_purpose at {DEFAULT_ACTIVITY_TIME}, selecting random activity")
                    current_month = current_time.month
                    logger.debug(f"Current month: {current_month}")

                    # Get activities for current month
                    month_activities = MAIN_ACTIVITIES_BY_MONTH.get(current_month, [])
                    if not month_activities:
                        month_activities = ["survey harvest", "seasonal conditions"]
                    logger.debug(f"Available activities for month {current_month}: {month_activities}")

                    # Select random activity
                    default_purpose = random.choice(month_activities)
                    logger.info(f"Selected random activity for user {user['user_id']}: {default_purpose}")
                else:
                    default_purpose = user['default_purpose']

                current_month = current_time.month
                current_year = current_time.year
                used_villages = set()
                month_activities = []

                # Get used villages and all activities for the month from new structure
                if year_str in activities and month_str in activities[year_str]:
                    month_activities = activities[year_str][month_str]
                    for activity in month_activities:
                        if activity.get('to_village'):
                            used_villages.add(activity['to_village'])

                villages = user.get('villages', [])
                available_villages = [v for v in villages if v not in used_villages]
                
                selection_pool = []
                if not available_villages:
                    # All villages visited once, reset the pool to all villages
                    selection_pool = villages[:] # Make a copy
                    
                    # Try to avoid picking the same village as the last recorded one
                    if month_activities:
                        # Sort by date to find the most recent activity
                        sorted_activities = self._sort_activities_by_date(month_activities)
                        last_village = sorted_activities[-1].get('to_village')
                        
                        # If there's a last village and more than one village in total
                        if last_village and len(selection_pool) > 1:
                            temp_pool = [v for v in selection_pool if v != last_village]
                            if temp_pool: # If the temp pool is not empty, use it
                                selection_pool = temp_pool
                else:
                    # Still have unique villages to visit this month
                    selection_pool = available_villages

                if not selection_pool:
                    logger.warning(f"Selection pool for user {user['user_id']} is empty. Skipping default activity.")
                    continue

                selected_village = random.choice(selection_pool)
                activity = {
                    'date': today_str,
                    'from': user['headquarters'] or 'HQ',
                    'to_village': selected_village,
                    'purpose': default_purpose
                }

                # Save using new structure
                users_collection.update_one(
                    {'user_id': user['user_id']},
                    {'$push': {f'activities.{year_str}.{month_str}': activity}}
                )

                # Delete the daily prompt message if it exists
                if user['user_id'] in self.daily_prompt_message_ids:
                    try:
                        msg_info = self.daily_prompt_message_ids[user['user_id']]
                        if isinstance(msg_info, dict):
                            message_id = msg_info['message_id']
                        else:
                            # Handle legacy format where only message_id was stored
                            message_id = msg_info

                        self.bot.delete_message(user['user_id'], message_id)

                        logger.info(f"🗑️ Deleted daily prompt message for user {user['user_id']}")
                        del self.daily_prompt_message_ids[user['user_id']]
                    except Exception as e:
                        logger.error(f"Error deleting daily prompt message for user {user['user_id']}: {e}")

                # Prepare message based on whether purpose was randomly selected
                purpose_explanation = (
                    "Since you havent added activity manually..."
                    if not user.get('default_purpose') else
                    f"Since you didn't record an activity by {DEFAULT_ACTIVITY_TIME}, I've added a default entry."
                )

                self.bot.send_message(
                    user['user_id'],
                    f"🤖 **Default Activity Recorded**\n\n"
                    f"{purpose_explanation}\n\n"
                    f"**Date:** {activity['date']}\n"
                    f"**From:** {activity['from']}\n"
                    f"**To:** {activity['to_village']}\n"
                    f"**Purpose:** {activity['purpose']}\n\n"
                    f"You can update this using /act command if needed.",
                    parse_mode='Markdown'
                )
            except Exception as e:
                logger.error(f"Error adding default activity for user {user['user_id']}: {e}")

    def schedule_daily_tasks(self):
        """Schedule daily tasks using the schedule library"""
        logger.info(f"📅 Setting up schedule: Daily prompt at {DAILY_PROMPT_TIME}, Default activity at {DEFAULT_ACTIVITY_TIME}")
        try:
            schedule.every().day.at(DAILY_PROMPT_TIME).do(self.daily_prompt)
            schedule.every().day.at(DEFAULT_ACTIVITY_TIME).do(self.default_activity_fallback)
            logger.info("✅ Schedule setup completed successfully")
        except Exception as e:
            logger.error(f"❌ Error setting up schedule: {e}")
            # Fallback to hardcoded times if variables fail
            logger.info("🔄 Using fallback hardcoded times")
            schedule.every().day.at("19:00").do(self.daily_prompt)
            schedule.every().day.at("20:00").do(self.default_activity_fallback)

        def run_schedule():
            logger.info("🔄 Schedule thread started - checking for pending tasks every 30 seconds")
            while True:
                try:
                    current_time = datetime.now(IST)
                    current_time_str = current_time.strftime('%H:%M')
                    # Ensure consistent format for schedule check messages
                    logger.debug(f"🕰️ Schedule check at {current_time.strftime('%H:%M:%S IST')}")

                    # Check if current time matches any scheduled times (ignoring seconds)
                    if current_time_str == DAILY_PROMPT_TIME:
                        logger.info(f"⏰ Daily prompt time matched: {current_time_str}")
                        self.daily_prompt()
                    elif current_time_str == DEFAULT_ACTIVITY_TIME:
                        logger.info(f"⏰ Default activity time matched: {current_time_str}")
                        self.default_activity_fallback()

                    time.sleep(30)
                except Exception as e:
                    logger.error(f"Error in schedule thread: {e}")
                    time.sleep(30)

        thread = threading.Thread(target=run_schedule, daemon=True)
        thread.start()

    def setup_handlers(self):
        """Setup bot handlers"""

        @self.bot.message_handler(commands=['start'])
        def start(message):
            self.start_command(message)

        @self.bot.message_handler(commands=['act'])
        def act(message):
            self.record_activity_command(message)

        @self.bot.message_handler(commands=['dnact'])
        def dnact(message):
            self.download_activities(message)

        @self.bot.message_handler(commands=['settings'])
        def settings(message):
            self.settings_command(message)

        @self.bot.message_handler(commands=['editact'])
        def editact(message):
            self.edit_activity_command(message)

        @self.bot.message_handler(content_types=['document'])
        def handle_docs(message):
            user_id = message.from_user.id
            if self.callback_data.get(user_id, {}).get('awaiting_holiday_upload'):
                self.handle_holiday_file_upload(message)
                self.callback_data[user_id].pop('awaiting_holiday_upload', None)
            elif self.callback_data.get(user_id, {}).get('awaiting_village_upload'):
                self.handle_file_upload(message)
                self.callback_data[user_id].pop('awaiting_village_upload', None)
            else:
                # User is not in a file upload flow, keep silent
                return

        @self.bot.message_handler(commands=['td'])
        def td_month(message):
            self.td_month_command(message)

        @self.bot.message_handler(commands=['ownerset'])
        def ownerset(message):
            self.owner_settings_command(message)

        @self.bot.message_handler(commands=['logs'])
        def send_logs(message):
            user_id = message.from_user.id
            if str(user_id) != str(OWNER_ID):
                self.bot.reply_to(message, "❌ You are not authorized to access logs.")
                return
            try:
                with open('logs.txt', 'rb') as f:
                    self.bot.send_document(message.chat.id, f, caption="📝 logs.txt")
            except FileNotFoundError:
                self.bot.reply_to(message, "No logs available yet.")
            except Exception as e:
                self.bot.reply_to(message, f"❌ Could not send logs.txt: {e}")

        @self.bot.message_handler(commands=['activities'])
        def activities_cmd(message):
            self.show_activities_years(message)

        @self.bot.callback_query_handler(func=lambda call: call.data.startswith(('activities_', 'delete_activity_', 'confirm_delete_')) or (call.data == 'cancel_selection' and hasattr(call.message, 'text') and '📅' in call.message.text))
        def activities_callback(call):
            user_id = call.from_user.id
            if call.data == 'cancel_selection':
                try:
                    self.bot.delete_message(call.message.chat.id, call.message.message_id)
                except Exception as e:
                    logger.error(f"Error deleting activities message: {e}")
            elif call.data.startswith('activities_year_'):
                year = call.data.split('_')[2]
                self.show_activities_months(call, year)
            elif call.data.startswith('activities_month_'):
                _, _, year, month = call.data.split('_')
                self.show_activities_dates(call, year, month)
            elif call.data.startswith('delete_activity_'):
                try:
                    _, _, year, month, index = call.data.split('_')
                    index = int(index)
                    user = users_collection.find_one({'user_id': user_id})
                    activities = user.get('activities', {})
                    month_activities = activities.get(year, {}).get(month, [])
                    if 0 <= index < len(month_activities):
                        month_activities = sorted(month_activities, key=lambda a: datetime.strptime(a['date'], '%d/%m/%Y'))
                        activity = month_activities[index]
                        keyboard = types.InlineKeyboardMarkup(row_width=2)
                        keyboard.add(
                            types.InlineKeyboardButton("✅ Yes", callback_data=f"confirm_delete_{year}_{month}_{index}"),
                            types.InlineKeyboardButton("❌ No", callback_data=f"activities_month_{year}_{month}")
                        )
                        msg = f"Are you sure you want to delete this activity?\n\n{activity['date']}: {activity.get('to_village','')} - {activity.get('purpose','')}\n"
                        self.bot.edit_message_text(msg, call.message.chat.id, call.message.message_id, reply_markup=keyboard)
                    else:
                        self.bot.answer_callback_query(call.id, "❌ Invalid activity index")
                except Exception as e:
                    logger.error(f"Error preparing delete confirmation: {e}")
                    self.bot.answer_callback_query(call.id, "❌ Error preparing delete confirmation")
            elif call.data.startswith('confirm_delete_'):
                try:
                    _, _, year, month, index = call.data.split('_')
                    index = int(index)
                    user = users_collection.find_one({'user_id': user_id})
                    activities = user.get('activities', {})
                    month_activities = activities.get(year, {}).get(month, [])
                    if 0 <= index < len(month_activities):
                        month_activities = sorted(month_activities, key=lambda a: datetime.strptime(a['date'], '%d/%m/%Y'))
                        deleted_activity = month_activities.pop(index)
                        activities[year][month] = month_activities
                        if not month_activities:
                            del activities[year][month]
                            if not activities[year]:
                                del activities[year]
                        users_collection.update_one(
                            {'user_id': user_id},
                            {'$set': {'activities': activities}}
                        )
                        self.show_activities_dates(call, year, month)
                        self.bot.answer_callback_query(call.id, f"✅ Activity deleted: {deleted_activity['date']}")
                    else:
                        self.bot.answer_callback_query(call.id, "❌ Invalid activity index")
                except Exception as e:
                    logger.error(f"Error deleting activity: {e}")
                    self.bot.answer_callback_query(call.id, "❌ Error deleting activity")

        self.callback_data = {}

        @self.bot.callback_query_handler(func=lambda call: True)
        def callback_query(call):
            user_id = call.from_user.id
            logger.info(f"Callback query from user {user_id}: {call.data}")

            # Remove the try: line here if it is not followed by except/finally
            # Unified cancel handling for both settings and selection
            if call.data in ('settings_cancel', 'cancel_selection'):
                logger.info(f"User {user_id} cancelled operation via {call.data}")
                if user_id in self.callback_data:
                    del self.callback_data[user_id]
                    logger.info(f"Cleaned up callback_data for user {user_id}")
                self.cancelled_users.add(user_id)  # Mark user as cancelled
                # Remove the inline keyboard (buttons) after cancelling
                try:
                    self.bot.edit_message_reply_markup(
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=None
                    )
                except Exception as e:
                    logger.error(f"Error removing inline keyboard after cancel for user {user_id}: {e}")
                self.bot.edit_message_text(
                    "❌ Operation cancelled.",
                    call.message.chat.id,
                    call.message.message_id
                )
                return
            # --- Store prompt message_id for input requests sent via edit_message_text ---
            if call.data == 'settings_activities':
                logger.info(f"settings_activities callback for user {user_id}")
                user = users_collection.find_one({'user_id': user_id})
                activities = user.get('custom_activities', []) if user else []
                activities_text = (
                    '\n'.join([f"{i+1}. {a}" for i, a in enumerate(activities)])
                    if activities else 'No custom activities defined.'
                )
                keyboard = types.InlineKeyboardMarkup()
                keyboard.add(
                    types.InlineKeyboardButton("➕ Add Activity", callback_data="settings_add_activity")
                )
                if activities:
                    keyboard.add(
                        types.InlineKeyboardButton("🗑️ Remove Activity", callback_data="settings_remove_activity")
                    )
                keyboard.add(
                    types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                )
                self.bot.edit_message_text(
                    f"📝 **Your Custom Activities:**\n\n{activities_text}\n\nYou can add or remove activities.",
                    call.message.chat.id,
                    call.message.message_id,
                    reply_markup=keyboard,
                    parse_mode='Markdown'
                )
                return
            if call.data == 'settings_remove_activity':
                logger.info(f"settings_remove_activity callback for user {user_id}")
                user = users_collection.find_one({'user_id': user_id})
                activities = user.get('custom_activities', []) if user else []
                if not activities:
                    self.bot.edit_message_text(
                        "❌ No activities to remove.",
                        call.message.chat.id,
                        call.message.message_id
                    )
                    return
                keyboard = types.InlineKeyboardMarkup()
                for i, a in enumerate(activities):
                    logger.info(f"Creating button for activity '{a}' at index {i}")
                    keyboard.add(
                        types.InlineKeyboardButton(f"🗑️ {a}", callback_data=f"remove_activity_idx_{i}")
                    )
                keyboard.add(
                    types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                )
                try:
                    self.bot.edit_message_text(
                        "🗑️ Select an activity to remove:",
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=keyboard
                    )
                except Exception as e:
                    logger.error(f"Failed to edit message for remove_activity for user {user_id}: {e}")
                    self.bot.send_message(
                        call.message.chat.id,
                        "🗑️ Select an activity to remove:",
                        reply_markup=keyboard
                    )
                return
            elif call.data.startswith('remove_activity_idx_'):
                logger.info(f"Processing remove_activity_idx_ for user {user_id}: callback_data='{call.data}'")
                try:
                    activity_idx = int(call.data.replace('remove_activity_idx_', ''))
                    logger.info(f"Parsed activity_idx: {activity_idx}")
                    user = users_collection.find_one({'user_id': user_id})
                    if not user:
                        logger.error(f"No user found for user_id {user_id}")
                        self.bot.edit_message_text(
                            "❌ User data not found. Please try again.",
                            call.message.chat.id,
                            call.message.message_id
                        )
                        return
                    activities = user.get('custom_activities', [])
                    logger.info(f"User activities: {activities}")
                    if not activities:
                        logger.warning(f"No activities to remove for user {user_id}")
                        self.bot.edit_message_text(
                            "❌ No activities to remove.",
                            call.message.chat.id,
                            call.message.message_id
                        )
                        return
                    if 0 <= activity_idx < len(activities):
                        activity = activities[activity_idx]
                        logger.info(f"Attempting to remove activity: '{activity}' at index {activity_idx}")
                        result = users_collection.update_one(
                            {'user_id': user_id},
                            {'$pull': {'custom_activities': activity}}
                        )
                        logger.info(f"Database update result: matched={result.matched_count}, modified={result.modified_count}")
                        if result.modified_count > 0:
                            logger.info(f"Successfully removed activity '{activity}' for user {user_id}")
                            # Refresh the activities UI after removal
                            user = users_collection.find_one({'user_id': user_id})
                            activities = user.get('custom_activities', []) if user else []
                            activities_text = (
                                '\n'.join([f"{i+1}. {a}" for i, a in enumerate(activities)])
                                if activities else 'No custom activities defined.'
                            )
                            keyboard = types.InlineKeyboardMarkup()
                            keyboard.add(
                                types.InlineKeyboardButton("➕ Add Activity", callback_data="settings_add_activity")
                            )
                            if activities:
                                keyboard.add(
                                    types.InlineKeyboardButton("🗑️ Remove Activity", callback_data="settings_remove_activity")
                                )
                            keyboard.add(
                                types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                            )
                            self.bot.edit_message_text(
                                f"✅ Activity removed: **{activity}**\n\n📝 **Your Custom Activities:**\n\n{activities_text}\n\nYou can add or remove activities.",
                                call.message.chat.id,
                                call.message.message_id,
                                reply_markup=keyboard,
                                parse_mode='Markdown'
                            )
                        else:
                            logger.error(f"Failed to remove activity '{activity}' for user {user_id}: No documents modified")
                            self.bot.edit_message_text(
                                f"❌ Failed to remove activity: **{activity}**. Please try again.",
                                call.message.chat.id,
                                call.message.message_id
                            )
                    else:
                        logger.warning(f"Invalid activity index {activity_idx} for user {user_id}, activity count: {len(activities)}")
                        self.bot.edit_message_text(
                            f"❌ Invalid activity selection (index {activity_idx}). Please try again.",
                            call.message.chat.id,
                            call.message.message_id
                        )
                except ValueError as ve:
                    logger.error(f"Invalid index format in callback_data '{call.data}' for user {user_id}: {ve}")
                    self.bot.edit_message_text(
                        f"❌ Invalid activity index. Please try again.",
                        call.message.chat.id,
                        call.message.message_id
                    )
                except Exception as e:
                    logger.error(f"Error removing activity at index for user {user_id}: {e}")
                    self.bot.edit_message_text(
                        f"❌ Error removing activity. Please try again.",
                        call.message.chat.id,
                        call.message.message_id
                    )
                return
            elif call.data == 'settings_add_activity':
                logger.info(f"settings_add_activity callback for user {user_id}")
                keyboard = types.InlineKeyboardMarkup()
                keyboard.add(
                    types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                )
                sent = self.bot.edit_message_text(
                    "📝 Please type the name of the activity to add:",
                    call.message.chat.id,
                    call.message.message_id,
                    reply_markup=keyboard
                )
                self.input_prompt_message[user_id] = sent.message_id
                self.bot.register_next_step_handler(
                    call.message,
                    lambda msg: self._handle_settings_add_activity_and_refresh(msg, call.message.chat.id)
                )
                return
            if call.data == 'settings_remove_village':
                logger.info(f"settings_remove_village callback for user {user_id}")
                user = users_collection.find_one({'user_id': user_id})
                villages = user.get('villages', []) if user else []
                logger.info(f"Showing remove village options for user {user_id}: villages={villages}")
                if not villages:
                    self.bot.edit_message_text(
                        "❌ No villages to remove.",
                        call.message.chat.id,
                        call.message.message_id
                    )
                    return
                keyboard = types.InlineKeyboardMarkup()
                for i, v in enumerate(villages):
                    logger.info(f"Creating button for village '{v}' at index {i}")
                    keyboard.add(
                        types.InlineKeyboardButton(f"🗑️ {v}", callback_data=f"remove_village_idx_{i}")
                    )
                keyboard.add(
                    types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                )
                try:
                    self.bot.edit_message_text(
                        "🗑️ Select a village to remove:",
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=keyboard
                    )
                except Exception as e:
                    logger.error(f"Failed to edit message for remove_village for user {user_id}: {e}")
                    self.bot.send_message(
                        call.message.chat.id,
                        "🗑️ Select a village to remove:",
                        reply_markup=keyboard
                    )
                return
            elif call.data.startswith('remove_village_idx_'):
                logger.info(f"Processing remove_village_idx_ for user {user_id}: callback_data='{call.data}'")
                try:
                    village_idx = int(call.data.replace('remove_village_idx_', ''))
                    logger.info(f"Parsed village_idx: {village_idx}")
                    user = users_collection.find_one({'user_id': user_id})
                    if not user:
                        logger.error(f"No user found for user_id {user_id}")
                        self.bot.edit_message_text(
                            "❌ User data not found. Please try again.",
                            call.message.chat.id,
                            call.message.message_id
                        )
                        return
                    villages = user.get('villages', [])
                    logger.info(f"User villages: {villages}")
                    if not villages:
                        logger.warning(f"No villages to remove for user {user_id}")
                        self.bot.edit_message_text(
                            "❌ No villages to remove.",
                            call.message.chat.id,
                            call.message.message_id
                        )
                        return
                    if 0 <= village_idx < len(villages):
                        match = villages[village_idx]
                        logger.info(f"Attempting to remove village: '{match}' at index {village_idx}")
                        result = users_collection.update_one(
                            {'user_id': user_id},
                            {'$pull': {'villages': match}}
                        )
                        logger.info(f"Database update result: matched={result.matched_count}, modified={result.modified_count}")
                        if result.modified_count > 0:
                            logger.info(f"Successfully removed village '{match}' for user {user_id}")
                            # Refresh the settings UI after removal
                            self._refresh_settings_ui(call.message.chat.id, user_id)
                        else:
                            logger.error(f"Failed to remove village '{match}' for user {user_id}: No documents modified")
                            self.bot.edit_message_text(
                                f"❌ Failed to remove village: **{match}**. Please try again.",
                                call.message.chat.id,
                                call.message.message_id
                            )
                    else:
                        logger.warning(f"Invalid village index {village_idx} for user {user_id}, village count: {len(villages)}")
                        self.bot.edit_message_text(
                            f"❌ Invalid village selection (index {village_idx}). Please try again.",
                            call.message.chat.id,
                            call.message.message_id
                        )
                except ValueError as ve:
                    logger.error(f"Invalid index format in callback_data '{call.data}' for user {user_id}: {ve}")
                    self.bot.edit_message_text(
                        f"❌ Invalid village index. Please try again.",
                        call.message.chat.id,
                        call.message.message_id
                    )
                except Exception as e:
                    logger.error(f"Error removing village at index for user {user_id}: {e}")
                    self.bot.edit_message_text(
                        f"❌ Error removing village. Please try again.",
                        call.message.chat.id,
                        call.message.message_id
                    )
                return
            elif call.data.startswith('purpose_'):
                temp_activity = self.callback_data.get(user_id, {})
                logger.info(f"Purpose callback - temp_activity for user {user_id}: {temp_activity}")

                if call.data == 'purpose_custom':
                    logger.info("Custom purpose selected, requesting text input")

                    # Delete the purpose selection message
                    try:
                        self.bot.delete_message(call.message.chat.id, call.message.message_id)
                        logger.debug(f"Deleted purpose selection message for user {user_id}")
                    except Exception as e:
                        logger.error(f"Failed to delete purpose selection message for user {user_id}: {e}")

                    # Send new message for custom purpose input
                    sent = self.bot.send_message(
                        call.message.chat.id,
                        "📝 Please type your custom purpose:"
                    )
                    self.bot.register_next_step_handler(
                        sent,
                        self.handle_purpose_selection,
                        temp_activity=temp_activity,
                        timeout=USER_TIMEOUT
                    )
                elif call.data.startswith('purpose_idx_'):
                    # Handle numbered activity selection
                    try:
                        idx = int(call.data.replace('purpose_idx_', ''))
                        user_activities = self.get_user_activities(user_id)

                        if 0 <= idx < len(user_activities):
                            purpose = user_activities[idx]
                            logger.info(f"Purpose selected by index {idx}: {purpose}")
                            temp_activity['purpose'] = purpose
                            temp_activity['user_id'] = user_id

                            # Delete the purpose selection message before saving
                            try:
                                self.bot.delete_message(call.message.chat.id, call.message.message_id)
                                logger.debug(f"Deleted purpose selection message for user {user_id}")
                            except Exception as e:
                                logger.error(f"Failed to delete purpose selection message for user {user_id}: {e}")

                            self.save_activity_callback(call, temp_activity)
                        else:
                            logger.error(f"Invalid purpose index {idx} for user {user_id}")
                            self.bot.answer_callback_query(call.id, "❌ Invalid selection. Please try again.")
                            return
                    except ValueError:
                        logger.error(f"Invalid purpose index format: {call.data}")
                        self.bot.answer_callback_query(call.id, "❌ Invalid selection. Please try again.")
                        return
                else:
                    # Handle legacy purpose selection (fallback)
                    purpose = call.data.replace('purpose_', '')
                    logger.info(f"Purpose selected (legacy): {purpose}")
                    temp_activity['purpose'] = purpose
                    temp_activity['user_id'] = user_id

                    # Delete the purpose selection message before saving
                    try:
                        self.bot.delete_message(call.message.chat.id, call.message.message_id)
                        logger.debug(f"Deleted purpose selection message for user {user_id}")
                    except Exception as e:
                        logger.error(f"Failed to delete purpose selection message for user {user_id}: {e}")

                    self.save_activity_callback(call, temp_activity)

                if user_id in self.callback_data:
                    del self.callback_data[user_id]
                    logger.info(f"Cleaned up callback_data for user {user_id}")
                    return

            elif call.data.startswith('settings_'):
                if call.data == 'settings_setrole':
                    logger.info(f"settings_setrole callback for user {user_id}")
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard.add(
                        types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                    )
                    self.bot.edit_message_text(
                        "👤 Please type your role/designation:",
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=keyboard
                    )
                    self.input_prompt_message[user_id] = call.message.message_id
                    self.bot.register_next_step_handler(
                        call.message,
                        self.handle_settings_setrole,
                        timeout=USER_TIMEOUT
                    )
                elif call.data == 'settings_sethq':
                    logger.info(f"settings_sethq callback for user {user_id}")
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard.add(
                        types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                    )
                    self.bot.edit_message_text(
                        "🏢 Please type your headquarters name:",
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=keyboard
                    )
                    self.input_prompt_message[user_id] = call.message.message_id
                    self.bot.register_next_step_handler(
                        call.message,
                        self.handle_settings_sethq,
                        timeout=USER_TIMEOUT
                    )
                    return
                elif call.data == 'settings_addvil':
                    logger.info(f"settings_addvil callback for user {user_id}")
                    user = users_collection.find_one({'user_id': user_id})
                    villages = user.get('villages', []) if user else []
                    villages_text = (
                        '\n'.join([f"{i+1}. {v}" for i, v in enumerate(villages)])
                        if villages else 'No villages added yet.'
                    )
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard.add(
                        types.InlineKeyboardButton("➕ Add Village", callback_data="settings_add_village")
                        )
                    if villages:
                        keyboard.add(
                            types.InlineKeyboardButton("🗑️ Remove Village", callback_data="settings_remove_village")
                        )
                    keyboard.add(
                        types.InlineKeyboardButton("📁 Upload File (Replace All)", callback_data="settings_upload_villages")
                    )
                    keyboard.add(
                        types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                    )
                    self.bot.edit_message_text(
                        f"🏘️ **Your Villages:**\n\n{villages_text}\n\nYou can add, remove, or upload a new list to replace all.",
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=keyboard,
                        parse_mode='Markdown'
                    )
                    return
                elif call.data == 'settings_add_village':
                    logger.info(f"settings_add_village callback for user {user_id}")
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard.add(
                        types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                    )
                    sent = self.bot.edit_message_text(
                        "🏘️ Please type the name of the village to add:",
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=keyboard
                    )
                    self.input_prompt_message[user_id] = sent.message_id
                    self.bot.register_next_step_handler(
                        call.message,
                        self.handle_settings_add_single_village,
                        timeout=USER_TIMEOUT
                    )
                    return
                elif call.data == 'settings_remove_village':
                    logger.info(f"settings_remove_village callback for user {user_id}")
                    user = users_collection.find_one({'user_id': user_id})
                    villages = user.get('villages', []) if user else []
                    logger.info(f"Showing remove village options for user {user_id}: villages={villages}")
                    if not villages:
                        self.bot.edit_message_text(
                            "❌ No villages to remove.",
                            call.message.chat.id,
                            call.message.message_id
                        )
                        return
                    keyboard = types.InlineKeyboardMarkup()
                    for i, v in enumerate(villages):
                        logger.info(f"Creating button for village '{v}' at index {i}")
                        keyboard.add(
                            types.InlineKeyboardButton(f"🗑️ {v}", callback_data=f"remove_village_idx_{i}")
                        )
                    keyboard.add(
                        types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                    )
                    try:
                        self.bot.edit_message_text(
                            "🗑️ Select a village to remove:",
                            call.message.chat.id,
                            call.message.message_id,
                            reply_markup=keyboard
                        )
                    except Exception as e:
                        logger.error(f"Failed to edit message for remove_village for user {user_id}: {e}")
                        self.bot.send_message(
                            call.message.chat.id,
                            "🗑️ Select a village to remove:",
                            reply_markup=keyboard
                        )
                    return
                elif call.data.startswith('remove_village_idx_'):
                    logger.info(f"Processing remove_village_idx_ for user {user_id}: callback_data='{call.data}'")
                    try:
                        village_idx = int(call.data.replace('remove_village_idx_', ''))
                        logger.info(f"Parsed village_idx: {village_idx}")
                        user = users_collection.find_one({'user_id': user_id})
                        if not user:
                            logger.error(f"No user found for user_id {user_id}")
                            self.bot.edit_message_text(
                                "❌ User data not found. Please try again.",
                                call.message.chat.id,
                                call.message.message_id
                            )
                            return
                        villages = user.get('villages', [])
                        logger.info(f"User villages: {villages}")
                        if not villages:
                            logger.warning(f"No villages to remove for user {user_id}")
                            self.bot.edit_message_text(
                                "❌ No villages to remove.",
                                call.message.chat.id,
                                call.message.message_id
                            )
                            return
                        if 0 <= village_idx < len(villages):
                            match = villages[village_idx]
                            logger.info(f"Attempting to remove village: '{match}' at index {village_idx}")
                            result = users_collection.update_one(
                                {'user_id': user_id},
                                {'$pull': {'villages': match}}
                            )
                            logger.info(f"Database update result: matched={result.matched_count}, modified={result.modified_count}")
                            if result.modified_count > 0:
                                logger.info(f"Successfully removed village '{match}' for user {user_id}")
                                # Refresh the settings UI after removal
                                self._refresh_settings_ui(call.message.chat.id, user_id)
                            else:
                                logger.error(f"Failed to remove village '{match}' for user {user_id}: No documents modified")
                                self.bot.edit_message_text(
                                    f"❌ Failed to remove village: **{match}**. Please try again.",
                                    call.message.chat.id,
                                    call.message.message_id
                                )
                        else:
                            logger.warning(f"Invalid village index {village_idx} for user {user_id}, village count: {len(villages)}")
                            self.bot.edit_message_text(
                                f"❌ Invalid village selection (index {village_idx}). Please try again.",
                                call.message.chat.id,
                                call.message.message_id
                            )
                    except ValueError as ve:
                        logger.error(f"Invalid index format in callback_data '{call.data}' for user {user_id}: {ve}")
                        self.bot.edit_message_text(
                            f"❌ Invalid village index. Please try again.",
                            call.message.chat.id,
                            call.message.message_id
                        )
                    except Exception as e:
                        logger.error(f"Error removing village at index for user {user_id}: {e}")
                        self.bot.edit_message_text(
                            f"❌ Error removing village. Please try again.",
                            call.message.chat.id,
                            call.message.message_id
                        )
                    return
                elif call.data == 'settings_upload_villages':
                    logger.info(f"settings_upload_villages callback for user {user_id}")
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard.add(
                        types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                    )
                    sent = self.bot.edit_message_text(
                        "📁 Please upload an Excel (.xlsx, .xls) or CSV (.csv) file with your villages. The file must have a column named 'Village'.",
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=keyboard
                    )
                    self.input_prompt_message[user_id] = sent.message_id
                    # The next document upload will be handled by handle_file_upload
                    self.callback_data[user_id] = {'awaiting_village_upload': True}
                    return
                elif call.data == 'settings_default_purpose':
                    logger.info(f"settings_default_purpose callback for user {user_id}")

                    # Check if user has a default purpose set
                    user = users_collection.find_one({'user_id': user_id})
                    current_purpose = user.get('default_purpose', None)

                    keyboard = types.InlineKeyboardMarkup()

                    if current_purpose:
                        # If user has a default purpose, show options to change or delete it
                        keyboard.add(
                            types.InlineKeyboardButton("✏️ Change Default Purpose", callback_data="settings_change_default_purpose")
                        )
                        keyboard.add(
                            types.InlineKeyboardButton("🗑️ Delete Default Purpose", callback_data="settings_delete_default_purpose")
                        )
                        keyboard.add(
                            types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
                        )

                        self.bot.edit_message_text(
                            f"🎯 Your current default purpose: *{current_purpose}*\n\nThis purpose is used for {DEFAULT_ACTIVITY_TIME} auto-entries.",
                            call.message.chat.id,
                            call.message.message_id,
                            reply_markup=keyboard,
                            parse_mode='Markdown'
                        )
                    else:
                        # If no default purpose is set, prompt to set one
                        keyboard.add(
                            types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
                        )

                        sent = self.bot.edit_message_text(
                            f"🎯 Please type your default purpose for {DEFAULT_ACTIVITY_TIME} auto-entries:",
                            call.message.chat.id,
                            call.message.message_id,
                            reply_markup=keyboard
                        )
                        self.input_prompt_message[user_id] = sent.message_id
                        self.bot.register_next_step_handler(
                            call.message,
                            self.handle_settings_default_purpose,
                            timeout=USER_TIMEOUT
                        )
                    return
                elif call.data == 'settings_change_default_purpose':
                    logger.info(f"settings_change_default_purpose callback for user {user_id}")
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard.add(
                        types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
                    )
                    sent = self.bot.edit_message_text(
                        f"🎯 Please type your new default purpose for {DEFAULT_ACTIVITY_TIME} auto-entries:",
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=keyboard
                    )
                    self.input_prompt_message[user_id] = sent.message_id
                    self.bot.register_next_step_handler(
                        call.message,
                        self.handle_settings_default_purpose,
                        timeout=USER_TIMEOUT,
                    )
                    return

            elif call.data.startswith('owner_'):
                if call.data == 'owner_set_prompt_time':
                    logger.info(f"owner_set_prompt_time callback for user {user_id}")
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard.add(
                        types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
                    )
                    sent = self.bot.edit_message_text(
                        "⏰ Please type the new daily prompt time (HH:MM):",
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=keyboard
                    )
                    self.input_prompt_message[user_id] = sent.message_id
                    self.bot.register_next_step_handler(
                        call.message,
                        self.handle_owner_set_prompt_time,
                        timeout=USER_TIMEOUT
                    )
                    return
                elif call.data == 'owner_set_fallback_time':
                    logger.info(f"owner_set_fallback_time callback for user {user_id}")
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard.add(
                        types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
                    )
                    sent = self.bot.edit_message_text(
                        "🤖 Please type the new default activity fallback time (HH:MM):",
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=keyboard
                    )
                    self.input_prompt_message[user_id] = sent.message_id
                    self.bot.register_next_step_handler(
                        call.message,
                        self.handle_owner_set_fallback_time,
                        timeout=USER_TIMEOUT,
                    )


                    return

                elif call.data == 'settings_delete_default_purpose':
                    logger.info(f"settings_delete_default_purpose callback for user {user_id}")
                    try:
                        # Get current default purpose
                        user = users_collection.find_one({'user_id': user_id})
                        current_purpose = user.get('default_purpose', None)

                        if not current_purpose:
                            self.bot.edit_message_text(
                                "❌ No default purpose is currently set.",
                                call.message.chat.id,
                                call.message.message_id
                            )
                            return

                        # Remove default purpose
                        result = users_collection.update_one(
                            {'user_id': user_id},
                            {'$unset': {'default_purpose': ""}}
                        )

                        if result.modified_count > 0:
                            logger.info(f"Successfully deleted default purpose for user {user_id}")
                            self.bot.edit_message_text(
                                f"✅ Default purpose deleted successfully. Auto-entries at {DEFAULT_ACTIVITY_TIME} will be disabled.",
                                call.message.chat.id,
                                call.message.message_id
                            )
                        else:
                            logger.error(f"Failed to delete default purpose for user {user_id}")
                            self.bot.edit_message_text(
                                "❌ Failed to delete default purpose. Please try again.",
                                call.message.chat.id,
                                call.message.message_id
                            )
                    except Exception as e:
                        logger.error(f"Error deleting default purpose for user {user_id}: {e}")
                        self.bot.edit_message_text(
                            "❌ An error occurred while deleting default purpose. Please try again.",
                            call.message.chat.id,
                            call.message.message_id
                        )
                    return
                elif call.data == 'settings_upload_holidays':
                    logger.info(f"settings_upload_holidays callback for user {user_id}")
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard.add(
                        types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
                    )
                    help_text = (
                        "📅 Please upload an Excel (.xlsx, .xls) or CSV (.csv) file with your public holidays.\n\n"
                        "The file must have two columns: 'Date' (A, DD/MM/YYYY) and 'Holiday' (B, description).\n\n"
                        "**Example:**\n"
                        "| Date        | Holiday            |\n"
                        "|-------------|--------------------|\n"
                        "| 15/08/2024  | Independence Day   |\n"
                        "| 02/10/2024  | Gandhi Jayanti     |\n"
                    )
                    sent = self.bot.edit_message_text(
                        help_text,
                        call.message.chat.id,
                        call.message.message_id,
                    disable_web_page_preview=True,
                        reply_markup=keyboard,
                        parse_mode='Markdown'
                    )
                    self.input_prompt_message[user_id] = sent.message_id
                    self.callback_data[user_id] = {'awaiting_holiday_upload': True}
                    return

            elif call.data.startswith('purpose_'):
                temp_activity = self.callback_data.get(user_id, {})
                logger.info(f"Purpose callback - temp_activity for user {user_id}: {temp_activity}")

                if call.data == 'purpose_custom':
                    logger.info("Custom purpose selected, requesting text input")

                    # Delete the purpose selection message
                    try:
                        self.bot.delete_message(call.message.chat.id, call.message.message_id)
                        logger.debug(f"Deleted purpose selection message for user {user_id}")
                    except Exception as e:
                        logger.error(f"Failed to delete purpose selection message for user {user_id}: {e}")

                    # Send new message for custom purpose input
                    sent = self.bot.send_message(
                        call.message.chat.id,
                        "📝 Please type your custom purpose:"
                    )
                    self.bot.register_next_step_handler(
                        sent,
                        self.handle_purpose_selection,
                        temp_activity=temp_activity,
                        timeout=USER_TIMEOUT
                    )
                elif call.data.startswith('purpose_idx_'):
                    # Handle numbered activity selection
                    try:
                        idx = int(call.data.replace('purpose_idx_', ''))
                        user_activities = self.get_user_activities(user_id)

                        if 0 <= idx < len(user_activities):
                            purpose = user_activities[idx]
                            logger.info(f"Purpose selected by index {idx}: {purpose}")
                            temp_activity['purpose'] = purpose
                            temp_activity['user_id'] = user_id

                            # Delete the purpose selection message before saving
                            try:
                                self.bot.delete_message(call.message.chat.id, call.message.message_id)
                                logger.debug(f"Deleted purpose selection message for user {user_id}")
                            except Exception as e:
                                logger.error(f"Failed to delete purpose selection message for user {user_id}: {e}")

                            self.save_activity_callback(call, temp_activity)
                        else:
                            logger.error(f"Invalid purpose index {idx} for user {user_id}")
                            self.bot.answer_callback_query(call.id, "❌ Invalid selection. Please try again.")
                            return
                    except ValueError:
                        logger.error(f"Invalid purpose index format: {call.data}")
                        self.bot.answer_callback_query(call.id, "❌ Invalid selection. Please try again.")
                        return
                elif call.data.startswith('purpose_'):
                    # Handle legacy purpose selection (fallback)
                    purpose = call.data.replace('purpose_', '')
                    logger.info(f"Purpose selected (legacy): {purpose}")
                    temp_activity['purpose'] = purpose
                    temp_activity['user_id'] = user_id

                    # Delete the purpose selection message before saving
                    try:
                        self.bot.delete_message(call.message.chat.id, call.message.message_id)
                        logger.debug(f"Deleted purpose selection message for user {user_id}")
                    except Exception as e:
                        logger.error(f"Failed to delete purpose selection message for user {user_id}: {e}")

                    self.save_activity_callback(call, temp_activity)

                if user_id in self.callback_data:
                    del self.callback_data[user_id]
                    logger.info(f"Cleaned up callback_data for user {user_id}")
                return

            # --- Village selection (normal, daily, or editact) ---
            if call.data.startswith('village_'):
                headquarters = users_collection.find_one({'user_id': user_id}).get('headquarters', 'HQ').title()
                if call.data == f'village_{headquarters}':
                    # User clicked headquarters button: no journey
                    date_str = self.callback_data.get(user_id, {}).get('date')
                    if not date_str:
                        date_str = datetime.now(IST).strftime('%d/%m/%Y')
                    temp_activity = {
                        'date': date_str,
                        'to_village': '',
                        'purpose': 'Attended office work',
                        'user_id': user_id
                    }
                    logger.info(f"User {user_id} selected headquarters for date {date_str}")
                    self.save_activity_callback(call, temp_activity)
                    if user_id in self.callback_data:
                        del self.callback_data[user_id]
                    return
                if call.data == 'village_manual':
                    logger.info(f"User {user_id} selected manual village entry")
                    date_str = self.callback_data.get(user_id, {}).get('date')
                    if not date_str:
                        date_str = datetime.now(IST).strftime('%d/%m/%Y')
                    temp_activity = {'date': date_str}
                    self.callback_data[user_id] = temp_activity

                    # Delete the village selection message
                    try:
                        self.bot.delete_message(call.message.chat.id, call.message.message_id)
                        logger.debug(f"Deleted village selection message for user {user_id}")
                    except Exception as e:
                        logger.error(f"Failed to delete village selection message for user {user_id}: {e}")

                    # Send new message for manual entry
                    sent = self.bot.send_message(
                        call.message.chat.id,
                        "✏️ Please type the village name:"
                    )
                    self.bot.register_next_step_handler(
                        sent,
                        self.handle_village_selection,
                        temp_activity=temp_activity,
                        timeout=USER_TIMEOUT
                    )
                    return
                village = call.data.replace('village_', '')
                logger.info(f"User {user_id} selected village: {village}")
                date_str = self.callback_data.get(user_id, {}).get('date')
                if not date_str:
                    date_str = datetime.now(IST).strftime('%d/%m/%Y')
                temp_activity = {
                    'to_village': village,
                    'date': date_str,
                    'user_id': user_id
                }
                self.callback_data[user_id] = temp_activity
                logger.info(f"Stored temp_activity for user {user_id}: {temp_activity}")

                # Delete the village selection message before showing purpose buttons
                try:
                    self.bot.delete_message(call.message.chat.id, call.message.message_id)
                    logger.debug(f"Deleted village selection message for user {user_id}")
                except Exception as e:
                    logger.error(f"Failed to delete village selection message for user {user_id}: {e}")

                # Extract month from date_str
                try:
                    selected_date = datetime.strptime(date_str, '%d/%m/%Y')
                    month = selected_date.month
                    logger.debug(f"Extracted month {month} from date {date_str} for purpose buttons")
                except Exception as e:
                    logger.error(f"Failed to extract month from date {date_str}: {e}")
                    month = datetime.now(IST).month
                    logger.debug(f"Using current month {month} for purpose buttons")

                self.show_purpose_buttons(call.message, user_id, month=month)
                return

            # --- Daily village selection ---
            elif call.data.startswith('daily_village_'):
                headquarters = users_collection.find_one({'user_id': user_id}).get('headquarters', 'HQ').title()
                if call.data == f'daily_village_{headquarters}':
                    # User clicked headquarters button: no journey
                    temp_activity = {
                        'date': datetime.now(IST).strftime('%d/%m/%Y'),
                        'to_village': '',
                        'purpose': 'Attended office work',
                        'user_id': user_id
                    }
                    logger.info(f"User {user_id} selected headquarters for daily activity")
                    self.save_activity_callback(call, temp_activity)
                    if user_id in self.callback_data:
                        del self.callback_data[user_id]
                    return
                if call.data == 'daily_village_manual':
                    logger.info(f"User {user_id} selected daily manual village entry")
                    temp_activity = {'date': datetime.now(IST).strftime('%d/%m/%Y')}
                    self.callback_data[user_id] = temp_activity

                    # Delete the daily village selection message
                    try:
                        self.bot.delete_message(call.message.chat.id, call.message.message_id)
                        logger.debug(f"Deleted daily village selection message for user {user_id}")
                    except Exception as e:
                        logger.error(f"Failed to delete daily village selection message for user {user_id}: {e}")

                    # Send new message for manual entry
                    sent = self.bot.send_message(
                        call.message.chat.id,
                        "✏️ Please type the village name:"
                    )
                    self.bot.register_next_step_handler(
                        sent,
                        self.handle_village_selection,
                        temp_activity=temp_activity,
                        timeout=USER_TIMEOUT
                    )
                    return
                village = call.data.replace('daily_village_', '')
                logger.info(f"User {user_id} selected daily village: {village}")
                temp_activity = {
                    'to_village': village,
                    'date': datetime.now(IST).strftime('%d/%m/%Y'),
                    'user_id': user_id
                }
                self.callback_data[user_id] = temp_activity
                logger.info(f"Stored daily temp_activity for user {user_id}: {temp_activity}")

                # Delete the daily village selection message before showing purpose buttons
                try:
                    self.bot.delete_message(call.message.chat.id, call.message.message_id)
                    logger.debug(f"Deleted daily village selection message for user {user_id}")
                except Exception as e:
                    logger.error(f"Failed to delete daily village selection message for user {user_id}: {e}")

                # Show purpose buttons for daily activity
                user_activities = self.get_user_activities(user_id)
                keyboard = types.InlineKeyboardMarkup()

                # Show numbered activities with numbered buttons in rows (side by side)
                row = []
                for i, purpose in enumerate(user_activities, 1):
                    row.append(types.InlineKeyboardButton(f"{i}", callback_data=f"daily_purpose_idx_{i-1}"))
                    # Add 5 buttons per row
                    if len(row) == 5:
                        keyboard.add(*row)
                        row = []

                # Add remaining buttons if any
                if row:
                    keyboard.add(*row)

                keyboard.add(
                    types.InlineKeyboardButton("📝 Manual Entry", callback_data="daily_purpose_custom")
                )
                keyboard.add(
                    types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
                )

                # Create numbered list text
                activities_text = '\n'.join([f"{i}. {purpose}" for i, purpose in enumerate(user_activities, 1)])
                message_text = f"🎯 **Select the purpose of visit:**\n\n{activities_text}\n\nClick the number button or use Manual Entry."

                sent = self.bot.send_message(
                    call.message.chat.id,
                    message_text,
                    reply_markup=keyboard,
                    parse_mode='Markdown'
                )
                return

            # --- Daily purpose selection ---
            elif call.data.startswith('daily_purpose_'):
                temp_activity = self.callback_data.get(user_id, {})
                logger.info(f"Daily purpose callback - temp_activity for user {user_id}: {temp_activity}")

                if call.data == 'daily_purpose_custom':
                    logger.info("Daily custom purpose selected, requesting text input")
                    self.bot.edit_message_text(
                        "📝 Please type your custom purpose:",
                        call.message.chat.id,
                        call.message.message_id
                    )
                    self.bot.register_next_step_handler(
                        call.message,
                        self.handle_purpose_selection,
                        temp_activity=temp_activity,
                        timeout=USER_TIMEOUT
                    )
                elif call.data.startswith('daily_purpose_idx_'):
                    # Handle numbered activity selection for daily
                    try:
                        idx = int(call.data.replace('daily_purpose_idx_', ''))
                        user_activities = self.get_user_activities(user_id)

                        if 0 <= idx < len(user_activities):
                            purpose = user_activities[idx]
                            logger.info(f"Daily purpose selected by index {idx}: {purpose}")
                            temp_activity['purpose'] = purpose
                            temp_activity['user_id'] = user_id
                            self.save_activity_callback(call, temp_activity)
                        else:
                            logger.error(f"Invalid daily purpose index {idx} for user {user_id}")
                            self.bot.answer_callback_query(call.id, "❌ Invalid selection. Please try again.")
                            return
                    except ValueError:
                        logger.error(f"Invalid daily purpose index format: {call.data}")
                        self.bot.answer_callback_query(call.id, "❌ Invalid selection. Please try again.")
                        return
                else:
                    # Handle legacy daily purpose selection (fallback)
                    purpose = call.data.replace('daily_purpose_', '')
                    logger.info(f"Daily purpose selected (legacy): {purpose}")
                    temp_activity['purpose'] = purpose
                    temp_activity['user_id'] = user_id
                    self.save_activity_callback(call, temp_activity)

                if user_id in self.callback_data:
                    del self.callback_data[user_id]
                    logger.info(f"Cleaned up callback_data for user {user_id}")
                return

    def edit_activity_command(self, message):
        """Handle /editact command for editing/adding activity for a specific date"""
        user_id = message.from_user.id
        user = users_collection.find_one({'user_id': user_id})

        logger.info(f"User {user_id} initiated /editact command")

        if not user or not user.get('villages'):
            logger.warning(f"User {user_id} has no villages configured")
            self.bot.reply_to(
                message, "❌ Please add villages first using /settings command."
            )
            return

        args = message.text.split()[1:] if len(message.text.split()) > 1 else []
        if args:
            date_str = args[0]
            try:
                datetime.strptime(date_str, '%d/%m/%Y')
            except ValueError:
                self.bot.reply_to(
                    message,
                    "❌ Invalid date format. Please use DD/MM/YYYY."
                )
                return
            logger.info(f"/editact date parsed: {date_str}")
            self._start_activity_flow_for_date(message, date_str)
        else:
            logger.info(f"/editact no date provided, prompting user {user_id} for date")
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(
                types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
            )
            sent = self.bot.send_message(
                message.chat.id,
                "📅 Please enter the date for the activity (DD/MM/YYYY):",
                reply_markup=keyboard
            )
            self.input_prompt_message[user_id] = sent.message_id
            self.bot.register_next_step_handler(
                message,
                self._editact_date_input_handler,
                timeout=USER_TIMEOUT
            )

    def _editact_date_input_handler(self, message, timeout=USER_TIMEOUT):
        user_id = message.from_user.id
        if user_id in self.cancelled_users:
            self.cancelled_users.remove(user_id)
            return
        date_str = message.text.strip()
        try:
            parsed_date = datetime.strptime(date_str, '%d/%m/%Y')
            month = parsed_date.month
            logger.debug(f"Successfully parsed date {date_str} to month {month} (type: {type(month)})")
            logger.debug(f"MAIN_ACTIVITIES_BY_MONTH[{month}]: {MAIN_ACTIVITIES_BY_MONTH.get(month)}")
        except ValueError:
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(
                types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
            )
            sent = self.bot.send_message(
                message.chat.id,
                "❌ Invalid date format. Please use DD/MM/YYYY:",
reply_markup=keyboard
            )
            self.input_prompt_message[user_id] = sent.message_id
            self.bot.register_next_step_handler(
                message,
                self._editact_date_input_handler,
                timeout=timeout
            )
            return
        logger.info(f"/editact user {user_id} provided date: {date_str}")
        # Delete the prompt message if present
        if user_id in self.input_prompt_message:
            try:
                self.bot.delete_message(message.chat.id, self.input_prompt_message[user_id])
            except Exception as e:
                logger.error(f"Error deleting prompt message for user {user_id}: {e}")
            del self.input_prompt_message[user_id]
        self._start_activity_flow_for_date(message, date_str)

    def _start_activity_flow_for_date(self, message, date_str):
        user_id = message.from_user.id
        user = users_collection.find_one({'user_id': user_id})
        logger.info(f"Starting activity flow for user {user_id} for date {date_str}")
        # Store the date in callback_data for this user
        self.callback_data[user_id] = {'date': date_str}
        self.show_village_buttons_for_date(message, user['villages'], date_str)

    def show_village_buttons_for_date(self, message, villages: List[str], date_str: str):
        """Show village selection buttons for a specific date, filtering out already visited villages for that month/year and notifying the user."""
        user_id = message.from_user.id
        user = users_collection.find_one({'user_id': user_id})

        # Ensure all village names are in proper case for display and comparison
        villages = [v.title() for v in villages]

        # Parse the month and year from the date_str
        try:
            selected_date = datetime.strptime(date_str, '%d/%m/%Y')
            selected_month = selected_date.month
            selected_year = selected_date.year
            logger.debug(f"Parsed date {date_str} in show_village_buttons_for_date: month={selected_month} (type: {type(selected_month)}), year={selected_year}")
            logger.debug(f"MAIN_ACTIVITIES_BY_MONTH[{selected_month}]: {MAIN_ACTIVITIES_BY_MONTH.get(selected_month)}")
        except Exception as e:
            logger.error(f"Invalid date_str in show_village_buttons_for_date: {date_str} - {e}")
            selected_month = None
            selected_year = None

        # Filter out already covered villages for the selected month/year
        covered_villages = set()
        if selected_month and selected_year:
            # Migrate to new structure if needed
            self.migrate_activities_structure(user_id)

            # Get activities from new structure
            activities = user.get('activities', {})
            year_str = str(selected_year)
            month_str = str(selected_month)

            if year_str in activities and month_str in activities[year_str]:
                for activity in activities[year_str][month_str]:
                    try:
                        activity_date = datetime.strptime(activity['date'], '%d/%m/%Y')
                        if activity_date.month == selected_month and activity_date.year == selected_year:
                            if activity.get('to_village'):
                                covered_villages.add(activity['to_village'].title())
                    except Exception:
                        continue
        available_villages = [v for v in villages if v not in covered_villages]

        keyboard = types.InlineKeyboardMarkup()
        headquarters = user.get('headquarters', 'HQ')
        hq_title = headquarters.title()
        keyboard.add(
            types.InlineKeyboardButton(
                f"🏢 {hq_title} (headquarters)", callback_data=f"village_{hq_title}"
            )
        )
        for i in range(0, len(available_villages), 2):
            row = []
            row.append(
                types.InlineKeyboardButton(
                    available_villages[i], callback_data=f"village_{available_villages[i]}"
                )
            )
            if i + 1 < len(available_villages):
                row.append(
                    types.InlineKeyboardButton(
                        available_villages[i + 1], callback_data=f"village_{available_villages[i + 1]}"
                    )
                )
            keyboard.add(*row)
        keyboard.add(
            types.InlineKeyboardButton("✏️ Manual Entry", callback_data="village_manual")
        )
        keyboard.add(
            types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
        )
        covered_text = (
            f"\n\n✅ Already covered this month: {len(covered_villages)} villages"
            if covered_villages else ""
        )
        sent = self.bot.send_message(
            message.chat.id,
            f"🏘️ Select the village you visited on {date_str}:{covered_text}",
            reply_markup=keyboard
        )
        # Track pending prompt for timeout
        self.pending_prompts[message.from_user.id] = {
            'message_id': sent.message_id,
            'chat_id': message.chat.id,
            'timeout_time': time.time() + USER_TIMEOUT,
            'type': 'button',
        }

    def td_month_command(self, message):
        """Handle /td <month_number> <year> command: send beautiful Excel of the month's tour diary"""
        user_id = message.from_user.id
        user = users_collection.find_one({'user_id': user_id})
        if not user:
            self.bot.reply_to(message, "❌ No activities found.")
            return
        args = message.text.split()[1:] if len(message.text.split()) > 1 else []
        if len(args) < 2:
            self.bot.reply_to(
                message,
                "❌ Please specify month and year. Usage: /td <month_number> <year> (e.g., /td 6 2024)"
            )
            return
        try:
            month_filter = int(args[0])
            year_filter = int(args[1])
            if month_filter < 1 or month_filter > 12:
                raise ValueError
        except ValueError:
            self.bot.reply_to(
                message,
                "❌ Invalid month or year. Usage: /td <month_number> <year> (e.g., /td 6 2024)"
            )
            return

        # Migrate to new structure if needed
        self.migrate_activities_structure(user_id)

        headquarters = user.get('headquarters', 'HQ')
        role = user.get('role')
        # Prepare a map of activities by date
        activities_by_date = {}
        activities = user.get('activities', {})
        year_str = str(year_filter)
        month_str = str(month_filter)

        if year_str in activities and month_str in activities[year_str]:
            for act in activities[year_str][month_str]:
                try:
                    dt = datetime.strptime(act['date'], '%d/%m/%Y')
                    activities_by_date[dt.date()] = act
                except Exception:
                    continue
        # Get all dates in the month
        num_days = calendar.monthrange(year_filter, month_filter)[1]
        all_dates = [datetime(year_filter, month_filter, day).date() for day in range(1, num_days+1)]
        # Get user public holidays (now a list of dicts)
        user_holidays = {}
        for h in user.get('public_holidays', []):
            try:
                dt = datetime.strptime(h['date'], '%d/%m/%Y').date()
                if dt.month == month_filter and dt.year == year_filter:
                    user_holidays[dt] = h['desc']
            except Exception:
                continue
        # Find all Sundays and second Saturday
        sundays = set()
        second_saturday = None
        for day in all_dates:
            if day.weekday() == 6:
                sundays.add(day)
        # Find second Saturday
        saturdays = [d for d in all_dates if d.weekday() == 5]
        if len(saturdays) >= 2:
            second_saturday = saturdays[1]
        # Prepare Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            logger.error("Failed to create Excel worksheet (ws is None)")
            self.bot.reply_to(message, "❌ Internal error creating Excel file.")
            return
        ws.title = 'Tour Diary'
        # Heading row (merged)
        first_name = getattr(message.from_user, 'first_name', '') or ''
        last_name = getattr(message.from_user, 'last_name', '') or ''
        username = getattr(message.from_user, 'username', '') or ''
        if first_name or last_name:
            user_name = (first_name + ' ' + last_name).strip()
        elif username:
            user_name = username.lstrip('@')
        else:
            user_name = 'User'
        month_name = calendar.month_name[month_filter]
        
        user_name_upper = user_name.upper()
        month_name_upper = month_name.upper()
        headquarters_upper = headquarters.upper()
        if role:
            heading_text = f"TOUR DIARY OF {user_name_upper}, {role}, {headquarters_upper} FOR THE MONTH OF {month_name_upper}-{year_filter}"
        else:
            heading_text = f"TOUR DIARY OF {user_name_upper}, {headquarters_upper} FOR THE MONTH OF {month_name_upper}-{year_filter}"
        ws.merge_cells('A1:D1')
        ws['A1'] = heading_text
        ws['A1'].font = Font(bold=True, size=13)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        # Headings in row 2
        headers = ['Date', 'From', 'To', 'Purpose of Journey']
        for i, h in enumerate(headers):
            ws.cell(row=2, column=i+1, value=h)
            ws.cell(row=2, column=i+1).font = Font(bold=True)
        # Data rows start from row 3
        tour_days = 0
        for day in all_dates:
            date_str = day.strftime('%d-%b-%Y')
            from_val = headquarters.title()
            if day in activities_by_date:
                act = activities_by_date[day]
                # Ensure From/To are proper case for consistency
                from_val = act.get('from', headquarters).title()
                to_val = act.get('to_village', '').title()
                purpose_val = act.get('purpose', '')
                
                # Custom logic to count tour days
                to_val_lower = to_val.lower()
                alpha_count = sum(1 for char in to_val if char.isalpha())
                
                if alpha_count >= 2 and 'leave' not in to_val_lower:
                    tour_days += 1
            elif day in sundays:
                to_val = ''
                purpose_val = 'Public holiday (Sunday)'
            elif second_saturday and day == second_saturday:
                to_val = ''
                purpose_val = 'Public holiday (Second Saturday)'
            elif day in user_holidays:
                to_val = ''
                purpose_val = f"Public holiday ({user_holidays[day]})"
            else:
                to_val = ''
                purpose_val = ''
            ws.append([
                date_str,
                from_val,
                to_val,
                purpose_val
            ])
        # Add a blank row after the table
        ws.append([''] * 4)
        summary_row = ws.max_row + 1
        ws.merge_cells(f'A{summary_row}:C{summary_row}')
        ws[f'A{summary_row}'] = f'No. of days toured in the month: {tour_days}'
        ws[f'A{summary_row}'].font = Font(bold=False)
        ws[f'A{summary_row}'].alignment = Alignment(horizontal='left', vertical='center')
        # Styling
        thin = Side(border_style="thin", color="000000")
        for row in ws.iter_rows(min_row=2, max_row=summary_row-1, min_col=1, max_col=4):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Autofit columns
        for col in ws.columns:
            max_length = 0
            col_index = col[0].column if hasattr(col[0], 'column') and col[0].column is not None else None
            if col_index is None:
                continue
            col_letter = get_column_letter(col_index)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 4
        # Set column A width to exactly 12
        ws.column_dimensions['A'].width = 12
        # Set column D ("Purpose of Journey") width to 55
        ws.column_dimensions['D'].width = 55
        # Autofit row heights
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            max_height = 1
            row_index = row[0].row if hasattr(row[0], 'row') and row[0].row is not None else None
            if row_index is None:
                continue
            for cell in row:
                if cell.value:
                    lines = str(cell.value).count('\n') + 1
                    max_height = max(max_height, lines * 15)
            ws.row_dimensions[row_index].height = max_height
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"TourDiary_{month_name}_{year_filter}.xlsx"
        self.bot.send_document(
            message.chat.id,
            document=(filename, output),
            caption=f"📋 Tour Diary for {month_name} {year_filter}"
        )

    def handle_settings_add_single_village(self, message, timeout=USER_TIMEOUT):
        user_id = message.from_user.id
        village = message.text.strip().title()
        # Validate village name: reject empty, commands, or invalid characters
        if not village or village.startswith('/') or not village.replace(' ', '').isalnum():
            self.bot.send_message(
                message.chat.id,
"❌ Invalid village name. Village names must be non-empty, not start with '/', and contain only letters, numbers, or spaces."
            )
            # Delete the prompt message if present
            if user_id in self.input_prompt_message:
                try:
                    self.bot.delete_message(message.chat.id, self.input_prompt_message[user_id])
                except Exception as e:
                    logger.error(f"Error deleting prompt message for user {user_id}: {e}")
                del self.input_prompt_message[user_id]
            # Re-show the village settings menu
            user = users_collection.find_one({'user_id': user_id})
            villages = user.get('villages', []) if user else []
            villages_text = (
                '\n'.join([f"{i+1}. {v}" for i, v in enumerate(villages)])
                if villages else 'No villages added yet.'
            )
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(
                types.InlineKeyboardButton("➕ Add Village", callback_data="settings_add_village")
                )
            if villages:
                keyboard.add(
                    types.InlineKeyboardButton("🗑️ Remove Village", callback_data="settings_remove_village")
                )
            keyboard.add(
                types.InlineKeyboardButton("📁 Upload File (Replace All)", callback_data="settings_upload_villages")
            )
            keyboard.add(
                types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
            )
            self.bot.send_message(
                message.chat.id,
                f"🏘️ **Your Villages:**\n\n{villages_text}\n\nYou can add, remove, or upload a new list to replace all.",
                reply_markup=keyboard,
                parse_mode='Markdown'
            )
            return
        users_collection.update_one(
            {'user_id': user_id},
            {'$addToSet': {'villages': village}},
            upsert=True
        )
        # Delete the prompt message if present
        if user_id in self.input_prompt_message:
            try:
                self.bot.delete_message(message.chat.id, self.input_prompt_message[user_id])
            except Exception as e:
                logger.error(f"Error deleting prompt message for user {user_id}: {e}")
            del self.input_prompt_message[user_id]
        # Refresh the village list UI after adding
        user = users_collection.find_one({'user_id': user_id})
        villages = user.get('villages', []) if user else []
        villages_text = (
            '\n'.join([f"{i+1}. {v}" for i, v in enumerate(villages)])
            if villages else 'No villages added yet.'
        )
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(
            types.InlineKeyboardButton("➕ Add Village", callback_data="settings_add_village")
        )
        if villages:
            keyboard.add(
                types.InlineKeyboardButton("🗑️ Remove Village", callback_data="settings_remove_village")
            )
        keyboard.add(
            types.InlineKeyboardButton("📁 Upload File (Replace All)", callback_data="settings_upload_villages")
        )
        keyboard.add(
            types.InlineKeyboardButton("❌ Cancel", callback_data="settings_cancel")
        )
        self.bot.send_message(
            message.chat.id,
            f"✅ Village added: **{village}**\n\n🏘️ **Your Villages:**\n\n{villages_text}\n\nYou can add, remove, or upload a new list to replace all.",
reply_markup=keyboard,
            parse_mode='Markdown'
        )

    def clean_invalid_villages(self):
        """One-time admin function to remove invalid village names from all users."""
        for user in users_collection.find({}):
            villages = user.get('villages', [])
            cleaned = [v for v in villages if v and not v.startswith('/') and v.replace(' ', '').isalnum()]
            if len(cleaned) != len(villages):
                users_collection.update_one({'user_id': user['user_id']}, {'$set': {'villages': cleaned}})
                logger.info(f"Cleaned villages for user {user['user_id']}: {cleaned}")

    def run(self):
        """Run the bot"""
        logger.info("Bot started successfully!")
        logger.info(f"Logging configured: logs.txt with max 6000 lines")
        try:
            self.bot.polling(none_stop=True)
        except Exception as e:
            logger.error(f"Bot polling error: {e}")
            time.sleep(15)
            self.run()

    def _refresh_settings_ui(self, chat_id, user_id):
        user = users_collection.find_one({'user_id': user_id})
        hq_status = (
            f"✅ {user.get('headquarters', 'Not set')}"
            if user.get('headquarters')
            else "❌ Not set"
        )
        villages_count = len(user.get('villages', []))
        custom_activities_count = len(user.get('custom_activities', []))
        default_purpose = user.get('default_purpose') or 'Not set'
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(
            types.InlineKeyboardButton("🏢 Set Headquarters", callback_data="settings_sethq")
        )
        keyboard.add(
            types.InlineKeyboardButton("🏘️ Add Villages", callback_data="settings_addvil")
        )
        keyboard.add(
            types.InlineKeyboardButton("📋 Manage Activities", callback_data="settings_activities")
        )
        keyboard.add(
            types.InlineKeyboardButton("🎯 Set Default Purpose", callback_data="settings_default_purpose")
        )
        keyboard.add(
            types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection")
        )
        settings_text = (
            f"⚙️ **Settings**\n\n"
            f"**Headquarters:** {hq_status}\n"
            f"**Role:** {role_status}\n"
            f"**Villages:** {villages_count} added\n"
            f"**Custom Activities:** {custom_activities_count} defined\n"
            f"**Default Purpose:** {default_purpose}\n"
            f"**Public Holidays:** {holidays_count} added\n\n"
            f"Select an option to configure:"
        )
        self.bot.send_message(
            chat_id,
            settings_text,
            reply_markup=keyboard,
            parse_mode='Markdown'
        )

    def _handle_settings_add_activity_and_refresh(self, message, chat_id):
        self.handle_settings_add_activity(message)
        user_id = message.from_user.id
        self._refresh_settings_ui(chat_id, user_id)

    def handle_holiday_file_upload(self, message):
        user_id = message.from_user.id
        file = message.document
        if not file:
            return
        file_name = file.file_name.lower()
        if not (file_name.endswith('.xlsx') or file_name.endswith('.xls') or file_name.endswith('.csv')):
            self.bot.reply_to(
                message,
                "❌ Please upload only Excel (.xlsx, .xls) or CSV (.csv) files."
            )
            return
        try:
            file_info = self.bot.get_file(file.file_id)
            downloaded_file = self.bot.download_file(file_info.file_path)
            file_data = BytesIO(downloaded_file)
            if file_name.endswith('.csv'):
                df = pd.read_csv(file_data)
            else:
                df = pd.read_excel(file_data)
            date_col = None
            desc_col = None
            for col in df.columns:
                if 'date' in col.lower():
                    date_col = col
                if 'holiday' in col.lower():
                    desc_col = col
            if date_col is None or desc_col is None:
                self.bot.reply_to(message, "❌ File must have both 'Date' and 'Holiday' columns.")
                return
            holidays = []
            skipped = []
            accepted_formats = ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d', '%Y/%m/%d']
            for i, row in df.iterrows():
                d_raw = row[date_col]
                desc = str(row[desc_col]).strip()
                if not d_raw or not desc:
                    reason = "Missing date or description"
                    skipped.append(f"Row {i+2}: '{d_raw}' - '{desc}' ({reason})")
                    continue
                # Accept datetime/date objects directly
                dt = None
                if hasattr(d_raw, 'year') and hasattr(d_raw, 'month') and hasattr(d_raw, 'day'):
                    try:
                        dt = d_raw
                    except Exception:
                        dt = None
                if dt is not None:
                    holidays.append({'date': dt.strftime('%d/%m/%Y'), 'desc': desc})
                    continue
                # Try string parsing
                d = str(d_raw).strip()
                parsed = False
                for fmt in accepted_formats:
                    try:
                        dt = datetime.strptime(d, fmt)
                        holidays.append({'date': dt.strftime('%d/%m/%Y'), 'desc': desc})
                        parsed = True
                        break
                    except Exception:
                        continue
                if not parsed:
                    reason = f"Invalid date format (got '{d}')"
                    skipped.append(f"Row {i+2}: '{d}' - '{desc}' ({reason})")
            if not holidays:
                error_msg = "❌ No valid holidays found in the file.\n\n"
                if skipped:
                    error_msg += "The following rows were skipped:\n" + '\n'.join(skipped[:10])
                    if len(skipped) > 10:
                        error_msg += f"\n...and {len(skipped)-10} more."
                else:
                    error_msg += "No valid rows detected."
                self.bot.reply_to(message, error_msg)
                return
            users_collection.update_one(
                {'user_id': user_id},
                {'$set': {'public_holidays': holidays}},
                upsert=True
            )
            reply_msg = (
                f"✅ Successfully added {len(holidays)} public holidays!\n\n"
                f"**Holidays added:** {', '.join([h['date'] + ' - ' + h['desc'] for h in holidays[:5]])}"
                + (f" and {len(holidays)-5} more..." if len(holidays) > 5 else "")
            )
            if skipped:
                reply_msg += f"\n\n⚠️ Skipped invalid rows: {', '.join(skipped[:5])}" + (f" and {len(skipped)-5} more..." if len(skipped) > 5 else "")
            self.bot.reply_to(message, reply_msg, parse_mode='Markdown')
            self._refresh_settings_ui(message.chat.id, user_id)
        except Exception as e:
            logger.error(f"Error processing holiday file for user {user_id}: {e}")
            self.bot.reply_to(message, f"❌ Error processing file: {str(e)}")

    def migrate_activities_structure(self, user_id):
        """Migrate flat activities list to nested year->month->list structure if needed."""
        user = users_collection.find_one({'user_id': user_id})
        if not user:
            return
        activities = user.get('activities', [])
        # If already in new structure, do nothing
        if isinstance(activities, dict):
            return
        # Otherwise, migrate
        new_activities = {}
        for act in activities:
            try:
                dt = datetime.strptime(act['date'], '%d/%m/%Y')
                year = str(dt.year)
                month = str(dt.month)
                if year not in new_activities:
                    new_activities[year] = {}
                if month not in new_activities[year]:
                    new_activities[year][month] = []
                new_activities[year][month].append(act)
            except Exception as e:
                logger.error(f"Migration: Skipping activity {act} for user {user_id}: {e}")
        users_collection.update_one({'user_id': user_id}, {'$set': {'activities': new_activities}})

    @staticmethod
    def _sort_activities_by_date(acts):
        return sorted(acts, key=lambda a: datetime.strptime(a['date'], '%d/%m/%Y'))

    def show_activities_years(self, message):
        user_id = message.from_user.id
        self.migrate_activities_structure(user_id)
        user = users_collection.find_one({'user_id': user_id})
        activities = user.get('activities', {})
        if not activities:
            self.bot.reply_to(message, "❌ No activities found.")
            return
        years = sorted(activities.keys())
        keyboard = types.InlineKeyboardMarkup()
        for y in years:
            keyboard.add(types.InlineKeyboardButton(y, callback_data=f"activities_year_{y}"))
        self.bot.send_message(message.chat.id, "📅 Select a year:", reply_markup=keyboard)

    def show_activities_months(self, call, year):
        user_id = call.from_user.id
        user = users_collection.find_one({'user_id': user_id})
        activities = user.get('activities', {})
        months = sorted(activities.get(year, {}).keys(), key=lambda m: int(m))
        keyboard = types.InlineKeyboardMarkup()
        for m in months:
            month_name = calendar.month_name[int(m)]
            keyboard.add(types.InlineKeyboardButton(month_name, callback_data=f"activities_month_{year}_{m}"))
        keyboard.add(types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection"))
        self.bot.edit_message_text(f"📅 Year: {year}\nSelect a month:", call.message.chat.id, call.message.message_id, reply_markup=keyboard)

    def show_activities_dates(self, call, year, month):
        user_id = call.from_user.id
        user = users_collection.find_one({'user_id': user_id})
        acts = user.get('activities', {}).get(year, {}).get(month, [])
        acts = sorted(acts, key=lambda a: datetime.strptime(a['date'], '%d/%m/%Y'))
        if not acts:
            self.bot.edit_message_text(f"❌ No activities for {calendar.month_name[int(month)]} {year}.", call.message.chat.id, call.message.message_id)
            return
        msg = f"📅 Activities for {calendar.month_name[int(month)]} {year}:\n\n"
        for i, act in enumerate(acts, 1):
            msg += f"{i}. {act['date']}: {act.get('to_village','')} - {act.get('purpose','')}\n"
        keyboard = types.InlineKeyboardMarkup(row_width=5)
        delete_buttons = []
        for i in range(1, len(acts) + 1):
            delete_buttons.append(types.InlineKeyboardButton(f"🗑️ {i}", callback_data=f"delete_activity_{year}_{month}_{i-1}"))
            if len(delete_buttons) == 5 or i == len(acts):
                keyboard.add(*delete_buttons)
                delete_buttons = []
        keyboard.add(types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_selection"))
        self.bot.edit_message_text(msg, call.message.chat.id, call.message.message_id, reply_markup=keyboard)

keep_alive()
if __name__ == '__main__':
    bot = TourDiaryBot()
    bot.run()
